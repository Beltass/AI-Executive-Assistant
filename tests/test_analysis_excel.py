"""Round-trip tests for the ``.xlsx`` deliverable.

THIS IS THE TEST THAT WAS MISSING. ``openpyxl`` was an undeclared dependency,
so on a clean install the entire Excel path — building a workbook AND reading
one — was dead, and nothing noticed. Importing ``openpyxl`` at module scope here
means that regression can never come back silently: if the dependency goes
missing, this file fails at collection.

Everything below builds a real workbook, writes it to disk, reopens it with
``openpyxl.load_workbook`` and asserts on what Excel would actually show:

* the sheet names, in Turkish, with the cover first;
* NATIVE chart objects (``BarChart`` / ``PieChart`` / ``LineChart``) — not
  pasted images — and their series references;
* conditional formatting colour-scale rules on the measure columns;
* frozen header panes and the autofilter on the raw data sheet;
* Turkish number formats (thousands, percent, seconds, date).

A screenshot in a spreadsheet is what "amatör duruyor" looks like; these
assertions are the difference.
"""

from __future__ import annotations

from datetime import datetime

import openpyxl
import pytest
from openpyxl.chart import BarChart, LineChart, PieChart

from ai_assistant.analysis import excel_report
from ai_assistant.analysis.analyzer import AnalysisRequest, analyze
from ai_assistant.analysis.dataset import DatasetError, load_csv, profile_table
from ai_assistant.analysis.excel_report import (
    FMT_DATE,
    FMT_INT,
    FMT_PCT,
    FMT_SEC,
    SHEET_CORRELATION,
    SHEET_COVER,
    SHEET_DATA,
    SHEET_KPI,
    SHEET_OUTLIERS,
    SHEET_SERIES,
    build_workbook,
)

FIXTURE = "tests/fixtures/cagri_merkezi.csv"
MOMENT = datetime(2026, 8, 1, 9, 30)


@pytest.fixture()
def workbook(tmp_path):
    """The full call-centre report, written to disk and read back."""
    result = analyze(load_csv(FIXTURE), now=MOMENT)
    path = build_workbook(result, str(tmp_path / "rapor.xlsx"), moment=MOMENT)
    assert path.endswith("rapor.xlsx")
    return openpyxl.load_workbook(path)


# --- structure ---------------------------------------------------------------


def test_the_file_opens_and_has_the_expected_turkish_sheets(workbook):
    names = workbook.sheetnames
    # The cover is written last but sits first, so its table of contents can
    # name the sheets that actually exist.
    assert names[0] == SHEET_COVER
    assert names[1] == SHEET_KPI
    assert names[2] == SHEET_DATA
    for expected in (SHEET_SERIES, SHEET_OUTLIERS, SHEET_CORRELATION):
        assert expected in names
    assert [n for n in names if n.startswith("Kırılım – ")], "no breakdown sheet"
    assert workbook.active.title == SHEET_COVER


def test_sheet_names_are_unique_and_within_excels_31_character_limit(workbook):
    names = workbook.sheetnames
    assert len(names) == len(set(names))
    assert all(len(name) <= 31 for name in names)
    assert not any(set(name) & set("[]:*?/\\") for name in names)


def test_the_cover_carries_the_facts_and_the_table_of_contents(workbook):
    sheet = workbook[SHEET_COVER]
    text = "\n".join(
        str(cell.value)
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Hazırlanma" in text
    assert "01.08.2026 09:30" in text
    assert "Veri kaynağı" in text
    assert "cagri_merkezi.csv" in text
    assert "112 satır × 12 sütun" in text
    assert "İçindekiler" in text
    assert "Çağrı merkezi operasyon verisi" in text
    assert sheet.sheet_view.showGridLines is False


def test_the_kpi_sheet_lists_findings_with_their_turkish_comment(workbook):
    sheet = workbook[SHEET_KPI]
    headers = [sheet.cell(row=4, column=c).value for c in range(1, 8)]
    assert headers == [
        "Gösterge",
        "Değer",
        "Karşılaştırma",
        "Fark",
        "Yön",
        "Durum",
        "Yorum",
    ]
    assert sheet.freeze_panes == "A5"

    labels = [sheet.cell(row=r, column=1).value for r in range(5, 20)]
    assert "Toplam çağrı" in labels
    assert "SLA tutturma" in labels
    # Every populated row carries a sentence, not just a number.
    for row in range(5, 5 + len([l for l in labels if l])):
        assert str(sheet.cell(row=row, column=7).value or "").strip()


# --- the raw data sheet ------------------------------------------------------


def test_the_data_sheet_is_frozen_filtered_and_complete(workbook):
    sheet = workbook[SHEET_DATA]
    assert [c.value for c in sheet[1]][:4] == ["Tarih", "Temsilci", "Kuyruk", "Vardiya"]
    # The header row stays put while you scroll 112 rows.
    assert sheet.freeze_panes == "A2"
    # …and every column is sortable/filterable from the header.
    assert sheet.auto_filter.ref == "A1:L113"
    assert sheet.max_row == 113


def test_data_cells_carry_turkish_number_formats(workbook):
    sheet = workbook[SHEET_DATA]
    headers = [c.value for c in sheet[1]]

    def fmt(name):
        return sheet.cell(row=2, column=headers.index(name) + 1).number_format

    assert fmt("Tarih") == FMT_DATE  # dd.mm.yyyy
    assert fmt("Çağrı Adedi") == FMT_INT  # #,##0
    assert fmt("AHT") == FMT_SEC  # #,##0" sn"
    assert fmt("SLA %") == FMT_PCT  # #,##0.0"%"
    assert fmt("Doluluk %") == FMT_PCT

    # The values are real typed values, not pre-formatted strings — Excel has
    # to be able to sort and chart them.
    assert sheet.cell(row=2, column=headers.index("Tarih") + 1).value == datetime(
        2026, 7, 1
    )
    assert sheet.cell(row=2, column=headers.index("Çağrı Adedi") + 1).value == 58
    assert sheet.cell(row=2, column=headers.index("AHT") + 1).value == 240


def test_booleans_are_written_as_turkish_words(tmp_path):
    data = profile_table(
        ["Vardiya", "Aktif", "Puan"],
        [["Sabah", "Evet", "5"], ["Akşam", "Hayır", "7"], ["Sabah", "Evet", "9"]],
    )
    path = build_workbook(analyze(data), str(tmp_path / "bool.xlsx"), moment=MOMENT)
    sheet = openpyxl.load_workbook(path)[SHEET_DATA]
    assert [sheet.cell(row=r, column=2).value for r in (2, 3, 4)] == [
        "Evet",
        "Hayır",
        "Evet",
    ]


# --- native charts -----------------------------------------------------------


def test_breakdown_sheets_carry_native_excel_charts(workbook):
    pivots = [name for name in workbook.sheetnames if name.startswith("Kırılım – ")]
    assert pivots

    charts_by_sheet = {name: workbook[name]._charts for name in pivots}
    for name, charts in charts_by_sheet.items():
        assert len(charts) == 1, name
        assert isinstance(charts[0], (BarChart, PieChart)), name
        # A real chart object knows its data range; an image would not.
        assert charts[0].series, name
        assert charts[0].title is not None, name

    # The first single-axis breakdown is small enough to read as a pie.
    first = workbook["Kırılım – Temsilci"]._charts[0]
    assert isinstance(first, PieChart)
    # The rest compare better as columns.
    others = [workbook[n]._charts[0] for n in pivots if n != "Kırılım – Temsilci"]
    assert others and all(isinstance(chart, BarChart) for chart in others)
    assert all(chart.type == "col" for chart in others)


def test_the_time_series_sheet_draws_a_line_per_series_plus_its_moving_average(
    workbook,
):
    sheet = workbook[SHEET_SERIES]
    charts = sheet._charts
    assert charts, "a dated call-centre file must produce trend charts"
    assert all(isinstance(chart, LineChart) for chart in charts)
    for chart in charts:
        # Two series: the value itself and the 3-period moving average.
        assert len(chart.series) == 2
        assert chart.x_axis.title is not None
        assert chart.y_axis.title is not None
        assert all(line.smooth is False for line in chart.series)


def test_a_bar_chart_points_at_the_cells_it_charts(workbook):
    chart = workbook["Kırılım – Kuyruk"]._charts[0]
    reference = str(chart.series[0].val.numRef.f)
    assert "Kırılım – Kuyruk" in reference or "Kırılım" in reference
    assert "$" in reference  # an absolute cell range, i.e. live data


# --- conditional formatting --------------------------------------------------


def test_measure_columns_get_a_colour_scale(workbook):
    sheet = workbook["Kırılım – Temsilci"]
    ranges = {str(rng.sqref): rng.rules for rng in sheet.conditional_formatting}
    assert ranges, "no conditional formatting on the breakdown sheet"

    rules = [rule for group in ranges.values() for rule in group]
    assert rules
    for rule in rules:
        assert rule.type == "colorScale"
        assert rule.colorScale is not None
        # min / percentile-50 / max — the three-stop scale.
        assert len(rule.colorScale.cfvo) == 3
        assert [c.type for c in rule.colorScale.cfvo] == ["min", "percentile", "max"]
        assert len(rule.colorScale.color) == 3


def test_the_correlation_sheet_scales_r_from_minus_one_to_plus_one(workbook):
    sheet = workbook[SHEET_CORRELATION]
    rules = [rule for rng in sheet.conditional_formatting for rule in rng.rules]
    assert len(rules) == 1
    scale = rules[0].colorScale
    assert [c.type for c in scale.cfvo] == ["num", "num", "num"]
    assert [float(c.val) for c in scale.cfvo] == [-1.0, 0.0, 1.0]


# --- the remaining sheets ----------------------------------------------------


def test_the_outlier_sheet_explains_the_method_and_the_band(workbook):
    sheet = workbook[SHEET_OUTLIERS]
    text = "\n".join(
        str(cell.value)
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "IQR" in text
    assert "Q1" in text and "Q3" in text
    assert sheet.freeze_panes == "A5"
    assert [sheet.cell(row=4, column=c).value for c in range(1, 7)] == [
        "Ölçüt",
        "Alt sınır",
        "Üst sınır",
        "Aykırı adet",
        "Oran",
        "Yorum",
    ]
    assert sheet.cell(row=5, column=5).number_format == FMT_PCT


def test_the_correlation_sheet_warns_that_correlation_is_not_causation(workbook):
    sheet = workbook[SHEET_CORRELATION]
    text = "\n".join(
        str(cell.value)
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "İlişki nedensellik" in text
    assert "Pearson" in text


def test_the_request_is_written_onto_the_cover(tmp_path):
    request = AnalysisRequest(
        dimensions=["Vardiya"],
        metrics=["sla"],
        last_days=7,
        question="Son bir haftada vardiya bazında SLA ne oldu?",
    )
    result = analyze(load_csv(FIXTURE), request, now=MOMENT)
    path = build_workbook(result, str(tmp_path / "istek.xlsx"), moment=MOMENT)
    sheet = openpyxl.load_workbook(path)[SHEET_COVER]
    text = "\n".join(
        str(cell.value)
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Sorulan soru" in text
    assert "Son bir haftada vardiya bazında SLA ne oldu?" in text
    assert "Analiz tarifi" in text
    assert "Son 1 hafta" in text


def test_notes_reach_the_cover_so_the_reader_knows_what_was_skipped(tmp_path):
    result = analyze(
        load_csv(FIXTURE),
        AnalysisRequest(metrics=["nps"]),
        now=MOMENT,
    )
    path = build_workbook(result, str(tmp_path / "not.xlsx"), moment=MOMENT)
    sheet = openpyxl.load_workbook(path)[SHEET_COVER]
    text = "\n".join(
        str(cell.value)
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Notlar" in text
    assert "bulunamayan gösterge" in text


# --- degradation -------------------------------------------------------------


def test_a_single_column_table_still_produces_a_readable_workbook(tmp_path):
    data = profile_table(["Tek"], [[str(i)] for i in range(1, 11)])
    path = build_workbook(analyze(data), str(tmp_path / "tek.xlsx"), moment=MOMENT)
    book = openpyxl.load_workbook(path)
    assert book.sheetnames[:3] == [SHEET_COVER, SHEET_KPI, SHEET_DATA]
    assert SHEET_SERIES not in book.sheetnames  # nothing to plot over time
    assert book[SHEET_DATA].max_row == 11


def test_an_all_text_table_produces_a_workbook_without_charts(tmp_path):
    data = profile_table(
        ["Şehir", "Not"], [["İstanbul", "iyi"], ["Ankara", "kötü"], ["İzmir", "iyi"]]
    )
    path = build_workbook(analyze(data), str(tmp_path / "metin.xlsx"), moment=MOMENT)
    book = openpyxl.load_workbook(path)
    assert SHEET_COVER in book.sheetnames
    assert book[SHEET_DATA].cell(row=2, column=1).value == "İstanbul"
    assert not any(
        sheet._charts for sheet in book.worksheets if sheet.title == SHEET_DATA
    )


def test_a_dimension_name_with_illegal_sheet_characters_is_cleaned(tmp_path):
    data = profile_table(
        ["Kanal/Ortam", "Değer"],
        [["A/B", "1"], ["C:D", "2"], ["A/B", "3"], ["C:D", "4"]],
    )
    path = build_workbook(analyze(data), str(tmp_path / "slash.xlsx"), moment=MOMENT)
    names = openpyxl.load_workbook(path).sheetnames
    assert "Kırılım – Kanal Ortam" in names
    assert not any("/" in name or ":" in name for name in names)


def test_the_workbook_creates_its_directory(tmp_path):
    target = tmp_path / "yeni" / "klasor" / "rapor.xlsx"
    result = analyze(profile_table(["A"], [["1"], ["2"]]))
    assert build_workbook(result, str(target), moment=MOMENT) == str(target)
    assert target.exists()


def test_an_unwritable_path_becomes_a_turkish_datasetError(tmp_path):
    """Saving onto a directory must not surface as a raw OSError."""
    directory = tmp_path / "burasi_bir_klasor.xlsx"
    directory.mkdir()
    result = analyze(profile_table(["A"], [["1"], ["2"]]))
    with pytest.raises(DatasetError) as excinfo:
        build_workbook(result, str(directory), moment=MOMENT)
    assert "kaydedilemedi" in str(excinfo.value)


def test_a_missing_openpyxl_is_reported_in_turkish(monkeypatch, tmp_path):
    """The safety net behind the declared dependency, exercised on purpose."""
    import builtins

    real_import = builtins.__import__

    def blocked(name, *args, **kwargs):
        if name == "openpyxl" or name.startswith("openpyxl."):
            raise ImportError("No module named 'openpyxl'")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", blocked)
    result = analyze(profile_table(["A"], [["1"], ["2"]]))
    with pytest.raises(DatasetError) as excinfo:
        build_workbook(result, str(tmp_path / "yok.xlsx"))
    message = str(excinfo.value)
    assert "openpyxl" in message
    assert "pip install openpyxl" in message


def test_openpyxl_is_actually_installed():
    """The declared dependency has to be present, or the engine is half dead.

    ``dataset.load_excel`` needs it to READ and ``excel_report`` needs it to
    WRITE — one missing package silently removes both halves of the Excel path.
    """
    assert excel_report._require_openpyxl() is openpyxl
    from openpyxl import load_workbook  # noqa: F401
