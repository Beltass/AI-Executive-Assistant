"""Tests for the analysis engine's front door: reading and profiling data.

This is where real-world breakage lives. The data a Turkish call centre exports
is not clean: ``1.234,56`` means one-thousand-two-hundred-thirty-four point
five-six, ``%85`` is a percentage, ``00:04:35`` is a duration that has to become
275 seconds, half the cells are ``N/A`` or blank, the headers carry
``ı/İ/ş/ğ/ö/ç/ü`` and the same file arrives in UTF-8 one week and CP1254 the
next. Getting any of that wrong does not raise — it silently produces a
confident, wrong report.

Just as important are the columns where parsing must NOT apply: an order code
``0012`` and a mobile number ``05321234567`` look numeric and are not, and
turning them into ``12.0`` and ``5321234567.0`` destroys them irreversibly.

Every failure path is asserted to arrive as a ``DatasetError`` carrying a plain
Turkish sentence — never a traceback. NO network: the Google Sheets path is
exercised with a stub service object.
"""

from __future__ import annotations

from datetime import date, datetime, timedelta

import pytest

from ai_assistant.analysis import dataset as ds
from ai_assistant.analysis.dataset import (
    KIND_BOOLEAN,
    KIND_CATEGORICAL,
    KIND_DATE,
    KIND_EMPTY,
    KIND_NUMERIC,
    KIND_TEXT,
    Dataset,
    DatasetError,
    detect_metric,
    load_any,
    load_csv,
    load_excel,
    load_google_sheet,
    load_text,
    parse_bool,
    parse_datetime,
    parse_number,
    profile_table,
    sheet_id_from_url,
)

FIXTURE = "tests/fixtures/cagri_merkezi.csv"


# --- Turkish number parsing --------------------------------------------------


@pytest.mark.parametrize(
    "raw, expected",
    [
        # The Turkish convention: dot groups thousands, comma is the decimal.
        ("1.234,56", 1234.56),
        ("1.234.567,89", 1234567.89),
        ("12,5", 12.5),
        ("0,75", 0.75),
        # The English convention has to keep working in the same file.
        ("1,234.56", 1234.56),
        ("1,234,567", 1234567.0),
        ("12.5", 12.5),
        # A lone separator with a three-digit tail is thousands, not decimals:
        # nobody writes 1 lira 234 kuruş.
        ("1.234", 1234.0),
        ("1.234.567", 1234567.0),
        # Percentages, written on either side, lose the sign and keep the value.
        ("%85", 85.0),
        ("85%", 85.0),
        ("%73,2", 73.2),
        ("% 99", 99.0),
        # Currency, spaces and accounting negatives.
        ("₺1.250,75", 1250.75),
        ("1.250,75 TL", 1250.75),
        ("$1,000.00", 1000.0),
        (" 1 234 ", 1234.0),
        ("(1.200)", -1200.0),
        ("-42,5", -42.5),
    ],
)
def test_parse_number_reads_turkish_and_english_notation(raw, expected):
    assert parse_number(raw) == pytest.approx(expected)


@pytest.mark.parametrize(
    "raw, seconds",
    [
        # A call-centre export writes AHT as a clock. hh:mm:ss.
        ("00:04:35", 275.0),
        ("01:00:00", 3600.0),
        ("02:13:07", 2 * 3600 + 13 * 60 + 7),
        # mm:ss — two fields in this file are minutes and seconds, never hours.
        ("04:35", 275.0),
        ("00:45", 45.0),
    ],
)
def test_parse_number_converts_durations_to_seconds(raw, seconds):
    assert parse_number(raw) == pytest.approx(seconds)


def test_parse_number_accepts_native_types():
    assert parse_number(42) == 42.0
    assert parse_number(3.5) == 3.5
    assert parse_number(timedelta(minutes=4, seconds=35)) == 275.0


@pytest.mark.parametrize(
    "raw",
    ["", "   ", None, "abc", "-", "N/A", "TR-0012", True, False, float("nan"), float("inf")],
)
def test_parse_number_refuses_what_is_not_a_number(raw):
    assert parse_number(raw) is None


# --- Turkish date parsing ----------------------------------------------------


@pytest.mark.parametrize(
    "raw, expected",
    [
        ("31.07.2026", datetime(2026, 7, 31)),
        ("31/07/2026", datetime(2026, 7, 31)),
        ("31-07-2026", datetime(2026, 7, 31)),
        ("2026-07-31", datetime(2026, 7, 31)),
        ("2026.07.31", datetime(2026, 7, 31)),
        ("31.07.2026 09:15", datetime(2026, 7, 31, 9, 15)),
        ("31.07.2026 09:15:30", datetime(2026, 7, 31, 9, 15, 30)),
        ("2026-07-31T14:30", datetime(2026, 7, 31, 14, 30)),
        ("2026-07", datetime(2026, 7, 1)),
        ("  2026-07-31  ", datetime(2026, 7, 31)),
    ],
)
def test_parse_datetime_reads_turkish_date_formats(raw, expected):
    assert parse_datetime(raw) == expected


def test_parse_datetime_accepts_date_objects_and_refuses_numbers():
    assert parse_datetime(date(2026, 7, 31)) == datetime(2026, 7, 31)
    assert parse_datetime(datetime(2026, 7, 31, 8)) == datetime(2026, 7, 31, 8)
    # A bare number is never a date here: 45000 is a call count, not an
    # Excel serial we may silently reinterpret.
    assert parse_datetime(45000) is None
    assert parse_datetime("dün") is None
    assert parse_datetime("") is None


def test_parse_bool_reads_turkish_words():
    for word in ("Evet", "VAR", "doğru", "aktif", "Açık"):
        assert parse_bool(word) is True
    for word in ("Hayır", "yok", "PASİF", "kapalı", "yanlış"):
        assert parse_bool(word) is False
    assert parse_bool("belki") is None
    assert parse_bool(None) is None


# --- what must NOT be parsed as a number -------------------------------------


def test_leading_zero_codes_stay_text_not_numbers():
    """``0012`` is a code. Parsing it to ``12.0`` deletes the code."""
    table = profile_table(
        ["Personel Kodu", "Puan"],
        [["0012", "5"], ["0034", "6"], ["0056", "7"], ["0078", "8"]],
    )
    code = table.column("Personel Kodu")
    assert code.kind == KIND_CATEGORICAL
    assert [row[0] for row in table.rows] == ["0012", "0034", "0056", "0078"]
    # And it must not be offered as something to sum or average.
    assert code not in table.numeric_columns()

    score = table.column("Puan")
    assert score.kind == KIND_NUMERIC


def test_turkish_mobile_numbers_keep_their_leading_zero():
    table = profile_table(
        ["Telefon"],
        [["05321234567"], ["05339876543"], ["05441112233"], ["05551234455"]],
    )
    assert table.column("Telefon").kind != KIND_NUMERIC
    assert table.rows[0][0] == "05321234567"


def test_a_genuine_number_starting_with_zero_is_still_a_number():
    """The guard is narrow: ``0``, ``0,75`` and ``0.5`` are values, not codes."""
    table = profile_table(
        ["Oran"], [["0"], ["0,75"], ["0,5"], ["1,25"], ["0"], ["2"]]
    )
    assert table.column("Oran").kind == KIND_NUMERIC
    assert table.rows[1][0] == pytest.approx(0.75)


def test_identifier_columns_are_not_read_as_call_centre_metrics():
    for header in (
        "Müşteri No",
        "Sipariş Kodu",
        "Kayıt ID",
        "Açıklama",
        "E-posta",
        "Ürün Adı",
        "Adres",
    ):
        assert detect_metric(header) is None, header


# --- call-centre metric detection (TR + EN) ----------------------------------


@pytest.mark.parametrize(
    "header, key",
    [
        ("AHT", "aht"),
        ("Ortalama Görüşme Süresi", "aht"),
        ("Handle Time", "aht"),
        ("Konuşma Süresi", "aht"),
        ("ASA", "asa"),
        ("Bekleme Süresi (sn)", "asa"),
        ("Answer Speed", "asa"),
        ("SLA %", "sla"),
        ("Servis Seviyesi", "sla"),
        ("Service Level", "sla"),
        ("Terk Edilen", "abandon"),
        ("Abandon Rate", "abandon"),
        ("Vazgeçen Çağrı", "abandon"),
        ("Doluluk %", "occupancy"),
        ("Occupancy", "occupancy"),
        ("Meşguliyet", "occupancy"),
        ("Memnuniyet", "csat"),
        ("CSAT", "csat"),
        ("Satisfaction Score", "csat"),
        ("FCR", "fcr"),
        ("İlk Temasta Çözüm", "fcr"),
        ("First Call Resolution", "fcr"),
        ("Çağrı Adedi", "calls"),
        ("Call Volume", "calls"),
        ("Cevaplanan", "answered"),
    ],
)
def test_detect_metric_finds_real_measures(header, key):
    hint = detect_metric(header)
    assert hint is not None, header
    assert hint.key == key
    assert hint.dimension is False


@pytest.mark.parametrize(
    "header, key",
    [
        ("Temsilci", "agent"),
        ("Müşteri Temsilcisi", "agent"),
        ("Agent", "agent"),
        ("Operatör", "agent"),
        ("Vardiya", "shift"),
        ("Shift", "shift"),
        ("Kuyruk", "queue"),
        ("Queue", "queue"),
        ("Kanal", "channel"),
        ("Takım", "team"),
        ("Tarih", "date"),
        ("Saat Dilimi", "interval"),
        ("Çağrı Sebebi", "reason"),
    ],
)
def test_detect_metric_finds_breakdown_axes(header, key):
    hint = detect_metric(header)
    assert hint is not None, header
    assert hint.key == key
    assert hint.dimension is True


@pytest.mark.parametrize(
    "header",
    ["Notlar", "Adres", "İl", "Şehir", "Renk", "Barkod", "IBAN", "", "   "],
)
def test_detect_metric_stays_quiet_on_unrelated_columns(header):
    assert detect_metric(header) is None


def test_detect_metric_survives_turkish_casing_and_separators():
    """``TEMSİLCİ`` casefolds to ``temsi̇lci`` — a dotted i the regex never sees."""
    assert detect_metric("TEMSİLCİ").key == "agent"
    assert detect_metric("temsilci_adı").key == "agent"
    assert detect_metric("ÇAĞRI-ADEDİ").key == "calls"
    assert detect_metric("ortalama.görüşme.süresi").key == "aht"
    assert detect_metric("İLK TEMAS").key == "fcr"


def test_more_specific_metric_wins_over_the_generic_one():
    """"Çağrı süresi" is AHT, not a call count — pattern order guarantees it."""
    assert detect_metric("Çağrı Süresi").key == "aht"
    assert detect_metric("Çağrı Adedi").key == "calls"


# --- profiling ---------------------------------------------------------------


def test_profile_reads_the_messy_column_without_losing_the_good_cells():
    table = profile_table(
        ["Tutar"],
        [["1.234,56"], ["  2.000  "], ["%85"], ["N/A"], ["-"], [""], ["3,5"], ["abc"], ["(500)"]],
    )
    column = table.column("Tutar")
    assert column.kind == KIND_NUMERIC
    # Rows that were blank end to end are dropped; "abc" survives as a row but
    # is counted as an invalid cell, not as a silent zero.
    assert column.invalid == 1
    assert column.count == 5
    assert column.minimum == pytest.approx(-500.0)
    assert column.maximum == pytest.approx(2000.0)
    assert [row[0] for row in table.rows] == [1234.56, 2000.0, 85.0, 3.5, None, -500.0]


def test_profile_types_each_column_by_its_own_content():
    table = profile_table(
        ["Tarih", "Temsilci", "Aktif", "Süre", "Boş"],
        [
            ["01.07.2026", "Ayşe Yılmaz", "Evet", "00:04:35", ""],
            ["02.07.2026", "Mehmet Demir", "Hayır", "00:05:00", ""],
            ["03.07.2026", "Ayşe Yılmaz", "Evet", "00:03:20", ""],
        ],
    )
    kinds = {c.name: c.kind for c in table.columns}
    assert kinds == {
        "Tarih": KIND_DATE,
        "Temsilci": KIND_CATEGORICAL,
        "Aktif": KIND_BOOLEAN,
        "Süre": KIND_NUMERIC,
        "Boş": KIND_EMPTY,
    }
    assert table.column("Süre").unit == "sn"
    assert table.rows[0][3] == 275.0
    assert table.column("Tarih").first == datetime(2026, 7, 1)
    assert table.column("Tarih").granularity == "gün"
    assert any("Boş" in note for note in table.notes)


def test_percent_unit_is_taken_from_the_cells_or_the_header():
    from_cells = profile_table(["Oran"], [["%80"], ["%90"], ["%70"]])
    assert from_cells.column("Oran").unit == "%"

    from_header = profile_table(["Doluluk %"], [["79,8"], ["86,6"], ["74,7"]])
    assert from_header.column("Doluluk %").unit == "%"


def test_long_free_text_becomes_text_not_a_category():
    rows = [[f"Müşteri {i} uzun uzun bir şikâyet metni yazdı ve bu satır kısa değil."] for i in range(12)]
    assert profile_table(["Not"], rows).column("Not").kind == KIND_TEXT


def test_headers_are_cleaned_deduplicated_and_ghost_columns_dropped():
    table = profile_table(
        ["Tarih", "Tarih", "  ", "  Çağrı   Adedi  ", None, ""],
        [["01.07.2026", "02.07.2026", "x", "5", None, None]],
    )
    # Trailing empty headers (Excel's ghost columns) disappear; a gap in the
    # middle is named; a duplicate is numbered.
    assert table.headers == ["Tarih", "Tarih (2)", "Sütun 3", "Çağrı Adedi"]


def test_column_lookup_ignores_case_and_separators():
    table = profile_table(["Çağrı Adedi"], [["5"], ["6"]])
    assert table.column("çağrı adedi") is not None
    assert table.column("Çağrı_Adedi") is not None
    assert table.column("yok") is None
    assert table.index_of("yok") == -1


def test_is_groupable_rejects_identity_columns():
    small = profile_table(["Vardiya"], [["Sabah"], ["Akşam"], ["Sabah"]])
    assert small.column("Vardiya").is_groupable is True

    identity = profile_table(["Kod"], [[f"MST-{i:04d}"] for i in range(120)])
    # 120 distinct values is an identity, not an axis you break a report down by.
    assert identity.column("Kod").is_groupable is False


def test_subset_keeps_the_declared_types_and_recomputes_the_statistics():
    table = profile_table(
        ["Vardiya", "SLA %"],
        [["Sabah", "%80"], ["Akşam", "%60"], ["Sabah", "%90"]],
    )
    morning = table.subset([row for row in table.rows if row[0] == "Sabah"])
    assert morning.row_count == 2
    profile = morning.column("SLA %")
    # The unit survives even though the values are now plain floats — re-running
    # inference on 80.0/90.0 would have thrown the "%" away.
    assert profile.unit == "%"
    assert profile.metric == "sla"
    assert profile.mean == pytest.approx(85.0)
    assert profile.minimum == pytest.approx(80.0)


def test_profile_dict_is_json_safe_and_stable():
    import json

    table = load_csv(FIXTURE)
    payload = table.profile()
    assert payload["profile_version"] == ds.PROFILE_VERSION
    assert payload["row_count"] == 112
    assert payload["column_count"] == 12
    assert payload["call_center_score"] == 1.0
    json.loads(table.to_json())  # must not raise
    assert table.profile() == payload  # same data, same profile


def test_content_hash_is_stable_and_sensitive():
    table = load_csv(FIXTURE)
    assert table.content_hash() == load_csv(FIXTURE).content_hash()

    changed = table.subset(table.rows[:-1])
    assert changed.content_hash() != table.content_hash()


def test_summary_tr_uses_turkish_thousands_separator():
    table = profile_table(["A"], [[str(i)] for i in range(1200)])
    assert table.summary_tr().startswith("1.200 satır × 1 sütun")


# --- the call-centre fixture end to end --------------------------------------


def test_fixture_is_recognised_as_call_centre_data():
    table = load_csv(FIXTURE)
    assert table.row_count == 112
    assert table.source_kind == "csv"
    assert table.call_center_score() == 1.0

    metrics = {c.name: c.metric for c in table.columns}
    assert metrics["Çağrı Adedi"] == "calls"
    assert metrics["AHT"] == "aht"
    assert metrics["SLA %"] == "sla"
    assert metrics["Terk Edilen"] == "abandon"
    assert metrics["Doluluk %"] == "occupancy"
    assert metrics["Memnuniyet"] == "csat"
    assert metrics["Temsilci"] == "agent"
    assert metrics["Vardiya"] == "shift"
    assert metrics["Kuyruk"] == "queue"

    assert [c.name for c in table.dimension_columns()] == ["Temsilci", "Kuyruk", "Vardiya"]


def test_fixture_first_row_is_parsed_cell_by_cell():
    table = load_csv(FIXTURE)
    first = table.rows[0]
    assert first[table.index_of("Tarih")] == datetime(2026, 7, 1)
    assert first[table.index_of("Temsilci")] == "Ayşe Yılmaz"
    assert first[table.index_of("Çağrı Adedi")] == 58.0
    assert first[table.index_of("AHT")] == 240.0          # "04:00"
    assert first[table.index_of("SLA %")] == pytest.approx(73.2)   # "%73,2"
    assert first[table.index_of("Doluluk %")] == pytest.approx(79.8)
    assert first[table.index_of("Memnuniyet")] == pytest.approx(3.63)

    assert table.column("AHT").unit == "sn"
    assert table.column("SLA %").unit == "%"
    assert table.column("Doluluk %").unit == "%"


def test_fixture_totals_match_the_source_file():
    table = load_csv(FIXTURE)
    assert table.column("Çağrı Adedi").total == pytest.approx(10382.0)
    assert table.column("Terk Edilen").total == pytest.approx(511.0)
    assert table.column("Tarih").first == datetime(2026, 7, 1)
    assert table.column("Tarih").last == datetime(2026, 7, 14)


# --- loaders -----------------------------------------------------------------


def test_load_text_sniffs_the_delimiter():
    tabbed = load_text("Ad\tPuan\nAyşe\t5\nMehmet\t7")
    assert tabbed.headers == ["Ad", "Puan"]
    assert tabbed.row_count == 2

    semicolons = load_text("Ad;Puan\nAyşe;5\nMehmet;7")
    assert semicolons.headers == ["Ad", "Puan"]

    commas = load_text("Ad,Puan\nAyşe,5\nMehmet,7")
    assert commas.headers == ["Ad", "Puan"]


def test_load_csv_reads_cp1254_when_utf8_fails(tmp_path):
    path = tmp_path / "eski.csv"
    path.write_bytes("Şehir;Çağrı\nİstanbul;120\nMuğla;45\n".encode("cp1254"))
    table = load_csv(str(path))
    assert table.headers == ["Şehir", "Çağrı"]
    assert table.rows[0][0] == "İstanbul"


def test_load_csv_strips_the_utf8_bom(tmp_path):
    path = tmp_path / "bom.csv"
    path.write_bytes("﻿Ad;Puan\nAyşe;5\nMehmet;7\n".encode("utf-8"))
    assert load_csv(str(path)).headers == ["Ad", "Puan"]


def test_load_any_routes_by_shape(tmp_path, monkeypatch):
    csv_path = tmp_path / "veri.csv"
    csv_path.write_text("Ad;Puan\nAyşe;5\nMehmet;7\n", encoding="utf-8")
    assert load_any(str(csv_path)).source_kind == "csv"
    assert load_any("Ad\tPuan\nAyşe\t5\nMehmet\t7").source_kind == "text"

    seen = {}

    def fake_sheet(url_or_id, sheet="", name=""):
        seen["url"] = url_or_id
        return profile_table(["A"], [["1"], ["2"]])

    monkeypatch.setattr(ds, "load_google_sheet", fake_sheet)
    load_any("https://docs.google.com/spreadsheets/d/" + "A" * 30 + "/edit#gid=0")
    assert "spreadsheets" in seen["url"]


def test_sheet_id_from_url():
    file_id = "1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgvE2upms"
    assert sheet_id_from_url(f"https://docs.google.com/spreadsheets/d/{file_id}/edit#gid=0") == file_id
    assert sheet_id_from_url(file_id) == file_id
    assert sheet_id_from_url("https://example.com/") == ""
    assert sheet_id_from_url("") == ""


# --- Excel loading -----------------------------------------------------------


def _write_workbook(path, rows, title="Veri"):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    for row in rows:
        sheet.append(row)
    workbook.save(str(path))
    return str(path)


def test_load_excel_reads_a_real_workbook(tmp_path):
    path = _write_workbook(
        tmp_path / "rapor.xlsx",
        [
            ["Tarih", "Temsilci", "Çağrı Adedi"],
            [datetime(2026, 7, 1), "Ayşe Yılmaz", 58],
            [datetime(2026, 7, 2), "Mehmet Demir", 66],
        ],
        title="Günlük",
    )
    table = load_excel(path)
    assert table.source_kind == "excel"
    assert table.name == "Günlük"
    assert table.headers == ["Tarih", "Temsilci", "Çağrı Adedi"]
    assert table.rows[0][0] == datetime(2026, 7, 1)
    assert table.rows[1][2] == 66.0


def test_load_excel_picks_the_busiest_sheet(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    cover = workbook.active
    cover.title = "Kapak"
    cover.append(["Bu bir kapak sayfasıdır"])
    data = workbook.create_sheet("Veriler")
    data.append(["Ad", "Puan"])
    for index in range(10):
        data.append([f"Kişi {index}", index])
    path = str(tmp_path / "iki.xlsx")
    workbook.save(path)

    table = load_excel(path)
    assert table.name == "Veriler"
    assert table.row_count == 10


def test_load_excel_reports_a_missing_sheet_in_turkish(tmp_path):
    path = _write_workbook(tmp_path / "tek.xlsx", [["A", "B"], [1, 2], [3, 4]], title="Sayfa1")
    with pytest.raises(DatasetError) as excinfo:
        load_excel(path, sheet="Olmayan")
    assert "bulunamadı" in str(excinfo.value)
    assert "Sayfa1" in str(excinfo.value)


# --- Google Sheets, with a stub service (no network) -------------------------


class _FakeValues:
    def __init__(self, values):
        self._values = values

    def get(self, **kwargs):
        self._kwargs = kwargs
        return self

    def execute(self):
        return {"values": self._values}


class _FakeSpreadsheets:
    def __init__(self, values, meta):
        self._values = values
        self._meta = meta

    def get(self, **kwargs):
        return self

    def execute(self):
        return self._meta

    def values(self):
        return _FakeValues(self._values)


class _FakeSheetsService:
    """Stands in for the Sheets API client. Nothing here touches the network."""

    def __init__(self, values, meta=None):
        self._sheets = _FakeSpreadsheets(
            values,
            meta or {"properties": {"title": "Temmuz"}, "sheets": [{"properties": {"title": "Günlük"}}]},
        )

    def spreadsheets(self):
        return self._sheets


def test_load_google_sheet_uses_the_injected_service():
    service = _FakeSheetsService(
        [["Tarih", "Çağrı Adedi"], ["01.07.2026", "58"], ["02.07.2026", "66"]]
    )
    table = load_google_sheet("https://docs.google.com/spreadsheets/d/" + "A" * 30, service=service)
    assert table.source_kind == "google_sheet"
    assert table.name == "Temmuz"
    assert table.row_count == 2
    assert table.rows[0][1] == 58.0


def test_load_google_sheet_reports_an_empty_sheet_in_turkish():
    service = _FakeSheetsService([["Tarih", "Çağrı"]])
    with pytest.raises(DatasetError) as excinfo:
        load_google_sheet("https://docs.google.com/spreadsheets/d/" + "A" * 30, service=service)
    assert "başlık dışında veri yok" in str(excinfo.value)


def test_load_google_sheet_translates_a_scope_error_into_advice():
    class _Boom:
        def spreadsheets(self):
            raise RuntimeError("HttpError 403: Request had insufficient authentication scopes.")

    with pytest.raises(DatasetError) as excinfo:
        load_google_sheet("https://docs.google.com/spreadsheets/d/" + "A" * 30, service=_Boom())
    message = str(excinfo.value)
    assert "spreadsheets.readonly" in message
    assert "google_auth" in message


def test_load_google_sheet_rejects_a_link_that_is_not_a_sheet():
    with pytest.raises(DatasetError) as excinfo:
        load_google_sheet("kısa")
    assert "tanınmadı" in str(excinfo.value)


def test_missing_google_credentials_produce_a_turkish_sentence(monkeypatch, tmp_path):
    for name in (
        "GOOGLE_CLIENT_ID",
        "GOOGLE_CLIENT_SECRET",
        "GOOGLE_CREDENTIALS_FILE",
        "GOOGLE_REFRESH_TOKEN",
    ):
        monkeypatch.delenv(name, raising=False)
    monkeypatch.setenv("GOOGLE_TOKEN_FILE", str(tmp_path / "yok.json"))

    with pytest.raises(DatasetError) as excinfo:
        load_google_sheet("https://docs.google.com/spreadsheets/d/" + "A" * 30)
    message = str(excinfo.value)
    assert "Google hesabı bağlı değil" in message
    assert "google_auth" in message


# --- graceful degradation: every failure is a Turkish sentence ---------------


def test_missing_file_is_a_turkish_error():
    with pytest.raises(DatasetError) as excinfo:
        load_csv("kesinlikle/olmayan/dosya.csv")
    assert "Dosya bulunamadı" in str(excinfo.value)


def test_empty_file_is_a_turkish_error(tmp_path):
    path = tmp_path / "bos.csv"
    path.write_text("", encoding="utf-8")
    with pytest.raises(DatasetError) as excinfo:
        load_csv(str(path))
    assert "Dosya boş" in str(excinfo.value)


def test_header_only_file_is_a_turkish_error(tmp_path):
    path = tmp_path / "sadece_baslik.csv"
    path.write_text("Ad;Puan\n", encoding="utf-8")
    with pytest.raises(DatasetError) as excinfo:
        load_csv(str(path))
    assert "başlık dışında satır yok" in str(excinfo.value)


def test_unreadable_file_is_a_turkish_error(tmp_path):
    """A directory where a file was expected must not surface as an OSError."""
    directory = tmp_path / "klasor"
    directory.mkdir()
    with pytest.raises(DatasetError) as excinfo:
        load_csv(str(directory))
    assert "okunamadı" in str(excinfo.value).casefold() or "boş" in str(excinfo.value)


def test_legacy_xls_is_refused_with_instructions(tmp_path):
    path = tmp_path / "eski.xls"
    path.write_bytes(b"\xd0\xcf\x11\xe0")
    with pytest.raises(DatasetError) as excinfo:
        load_excel(str(path))
    assert ".xlsx" in str(excinfo.value)


def test_corrupt_workbook_is_a_turkish_error(tmp_path):
    path = tmp_path / "bozuk.xlsx"
    path.write_bytes(b"bu bir zip degil")
    with pytest.raises(DatasetError) as excinfo:
        load_excel(str(path))
    assert "açılamadı" in str(excinfo.value)


def test_empty_paste_and_one_line_paste_are_turkish_errors():
    with pytest.raises(DatasetError) as excinfo:
        load_text("   ")
    assert "boş" in str(excinfo.value)

    with pytest.raises(DatasetError) as excinfo:
        load_text("sadece bir başlık satırı")
    assert "en az iki satır" in str(excinfo.value)


def test_unknown_source_is_a_turkish_error():
    with pytest.raises(DatasetError) as excinfo:
        load_any("")
    assert "veri kaynağı verilmedi" in str(excinfo.value)

    with pytest.raises(DatasetError) as excinfo:
        load_any("https://example.com/rapor.html")
    assert "Google E-Tablo değil" in str(excinfo.value)

    with pytest.raises(DatasetError) as excinfo:
        load_any("ne dosya ne bağlantı")
    assert "tanınmadı" in str(excinfo.value)


def test_a_table_with_no_body_rows_is_a_turkish_error():
    with pytest.raises(DatasetError) as excinfo:
        profile_table(["Ad", "Puan"], [])
    assert "hiç satır yok" in str(excinfo.value)


def test_a_table_with_no_headers_is_a_turkish_error():
    with pytest.raises(DatasetError) as excinfo:
        profile_table([], [["a"]])
    assert "sütun başlığı bulunamadı" in str(excinfo.value)


def test_ragged_rows_are_padded_not_dropped():
    table = profile_table(["A", "B", "C"], [["1", "2", "3"], ["4"], ["5", "6"]])
    assert table.row_count == 3
    assert table.rows[1] == [4.0, None, None]
