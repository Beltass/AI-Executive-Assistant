"""Veri alma ve profilleme — motorun giriş kapısı.

Kullanıcı elindeki veriyi nasıl saklıyorsa öyle verir: bir ``.xlsx`` dosyası, bir
``.csv``, bir Google E-Tablo bağlantısı ya da doğrudan yapıştırılmış bir tablo.
Bu modül dördünü de aynı :class:`Dataset` nesnesine indirger, sonra her sütunu
profiller: tipi, boşluk oranı, benzersiz değer sayısı, sayısal sütunlar için
özet istatistikler, kategorik sütunlar için en sık değerler, tarih sütunları için
saptanan periyot.

Profil JSON'a çevrilebilir ve KARARLIDIR: aynı veri her zaman aynı profili verir,
çünkü hem Excel raporu hem web raporu hem de canlı senkronizasyon aynı sözlüğün
üstünde çalışır.

TÜRKÇE VERİYLE ÇALIŞIR. Sayı ayrıştırıcı hem ``1.234,56`` hem ``1,234.56``
biçimini, ``%85`` ve ``85%`` yazımını, ``₺``/``TL`` işaretini ve ``00:04:35``
süre biçimini tanır. Tarih ayrıştırıcı ``31.07.2026``, ``31/07/2026``,
``2026-07-31`` ve saatli türevlerini okur.

ÇAĞRI MERKEZİ SEZGİSİ. Sütun adları hem Türkçe hem İngilizce anahtar sözcüklerle
eşlenir (``çağrı/call``, ``bekleme/wait``, ``AHT``, ``ASA``, ``SLA``,
``terk/abandon``, ``doluluk/occupancy``, ``memnuniyet/CSAT``, ``FCR``,
``vardiya/shift``, ``temsilci/agent``). Eşleşen sütun, hangi analizin varsayılan
olarak çalışacağını belirler — kullanıcı tek bir ayar yapmadan doğru raporu alır.

Hata yönetimi tektir: her sorun :class:`DatasetError` ile ve TÜRKÇE bir cümleyle
yükselir. Çağıran taraf tek bir ``except DatasetError`` ile kullanıcıya
gösterilebilir bir mesaj elde eder.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import math
import os
import re
import statistics
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

logger = logging.getLogger(__name__)

#: Profil sözlüğünün sürümü. Canlı senkronizasyon eski sürümleri tanıyabilsin
#: diye her belgeye yazılır.
PROFILE_VERSION = 1

#: Kategorik bir sütun için saklanan en sık değer sayısı.
TOP_VALUES = 10

#: Tip çıkarımı için örneklenecek en fazla satır. Milyon satırlık bir dosyada
#: tip kararı ilk birkaç bin satırda zaten kesinleşir.
TYPE_SAMPLE = 2000

#: Bir sütunun tipe ait sayılması için gereken başarı oranı.
TYPE_THRESHOLD = 0.7

#: Bellekte tutulan en fazla satır. Bunun üstü kırpılır ve rapora not düşülür —
#: 200 bin satırlık bir dışa aktarım yüzünden süreç ölmemeli.
MAX_ROWS = 200_000

KIND_NUMERIC = "numeric"
KIND_DATE = "date"
KIND_CATEGORICAL = "categorical"
KIND_TEXT = "text"
KIND_BOOLEAN = "boolean"
KIND_EMPTY = "empty"

KIND_LABELS = {
    KIND_NUMERIC: "sayısal",
    KIND_DATE: "tarih",
    KIND_CATEGORICAL: "kategorik",
    KIND_TEXT: "metin",
    KIND_BOOLEAN: "evet/hayır",
    KIND_EMPTY: "boş",
}


class DatasetError(RuntimeError):
    """Veri okunamadığında TÜRKÇE mesajla yükselen tek hata tipi."""


# --- küçük ayrıştırıcılar ----------------------------------------------------

_CURRENCY = "₺$€£"
#: Para birimi SÖZCÜĞÜ önce denenir: aksi halde "1.250,75 TL" yazımındaki
#: boşluk, boşluk kuralına takılır, "TL" yalnız kalır ve sayı hiç
#: ayrıştırılamazdı — oysa Türkçe bir dışa aktarımda tutar tam olarak
#: böyle yazılır.
_NUM_STRIP = re.compile(
    r"(?:^|(?<=\d))\s*(?:tl|try|usd|eur)\b|[\s " + _CURRENCY + r"]",
    re.IGNORECASE,
)
_DURATION_RE = re.compile(r"^(\d{1,3}):([0-5]?\d)(?::([0-5]?\d))?$")
_PERCENT_RE = re.compile(r"^\s*%\s*|\s*%\s*$")

#: Baştaki sıfır bir SAYININ değil bir KODUN parçasıdır: ``0012`` personel
#: kodu, ``05321234567`` telefon, ``0045`` sipariş numarası. Bunları sayıya
#: çevirmek baştaki sıfırı — yani kodun kendisini — siler, üstelik ortaya
#: toplanabilir görünen anlamsız bir "ölçüt" çıkarır.
_LEADING_ZERO_CODE = re.compile(r"^0\d+$")

_TRUE_WORDS = {"evet", "var", "true", "yes", "doğru", "dogru", "e", "y", "1", "aktif", "açık", "acik"}
_FALSE_WORDS = {"hayır", "hayir", "yok", "false", "no", "yanlış", "yanlis", "h", "n", "0", "pasif", "kapalı", "kapali"}

_DATE_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M",
    "%Y-%m-%d",
    "%d.%m.%Y %H:%M:%S",
    "%d.%m.%Y %H:%M",
    "%d.%m.%Y",
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%Y/%m/%d",
    "%m/%d/%Y %H:%M",
    "%m/%d/%Y",
    "%Y.%m.%d",
    "%Y-%m",
)


def parse_number(value: Any) -> Optional[float]:
    """Bir hücreyi sayıya çevirir; çeviremezse ``None``.

    Türkçe ve İngilizce basamak ayırıcılarının ikisini de kabul eder. ``%85``
    yüzde işareti düşülerek ``85`` olur (yüzde birimi sütun profilinde tutulur),
    ``00:04:35`` biçimindeki süre SANİYEye çevrilir — çağrı merkezi dışa
    aktarımlarında AHT neredeyse her zaman bu biçimdedir.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return None if isinstance(value, float) and not math.isfinite(value) else float(value)
    if isinstance(value, timedelta):
        return float(value.total_seconds())
    text = str(value).strip()
    if not text:
        return None

    text = _PERCENT_RE.sub("", text).strip()
    match = _DURATION_RE.match(text)
    if match:
        hours, minutes, seconds = match.group(1), match.group(2), match.group(3)
        if seconds is None:
            # "04:35" bir çağrı merkezi dosyasında saat değil dakika:saniyedir.
            return float(int(hours) * 60 + int(minutes))
        return float(int(hours) * 3600 + int(minutes) * 60 + int(seconds))

    text = _NUM_STRIP.sub("", text)
    if not text or not re.search(r"\d", text):
        return None
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    if not re.fullmatch(r"[+-]?[\d.,]*\d[\d.,]*", text):
        return None

    has_dot = "." in text
    has_comma = "," in text
    if has_dot and has_comma:
        decimal = "," if text.rfind(",") > text.rfind(".") else "."
        thousands = "." if decimal == "," else ","
        text = text.replace(thousands, "").replace(decimal, ".")
    elif has_comma:
        text = _single_separator(text, ",")
    elif has_dot:
        text = _single_separator(text, ".")

    try:
        number = float(text)
    except ValueError:
        return None
    if not math.isfinite(number):
        return None
    return -number if negative else number


def _single_separator(text: str, sep: str) -> str:
    """Tek tür ayırıcı taşıyan bir sayıda o ayırıcının rolünü kararlaştırır."""
    parts = text.replace("+", "").lstrip("-").split(sep)
    # "1.234.567" / "1,234,567" → binlik; "12.5" / "12,5" → ondalık.
    if len(parts) > 2 and all(len(p) == 3 for p in parts[1:]):
        return text.replace(sep, "")
    if len(parts) == 2 and len(parts[1]) == 3 and len(parts[0]) <= 3 and parts[0].isdigit():
        # 1.234 ikircikli: binlik ayırıcı kabul edilir (nokta) ama virgülde
        # Türkçe ondalık daha olası değil — üç haneli kuruş yazılmaz.
        return text.replace(sep, "") if sep == "." else text.replace(sep, ".")
    return text.replace(sep, ".")


def parse_datetime(value: Any) -> Optional[datetime]:
    """Bir hücreyi ``datetime``a çevirir; çeviremezse ``None``."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, time):
        return None
    if isinstance(value, (int, float)):
        return None
    text = str(value).strip()
    if not text or not re.search(r"\d", text):
        return None
    cleaned = text.replace("Z", "").strip()
    if len(cleaned) > 32:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(cleaned, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(cleaned)
    except ValueError:
        return None


def _tr_casefold(text: str) -> str:
    """Türkçe kurallı küçültme.

    Python'un ``casefold()``ı ASCII kuralıyla çalışır: ``"AÇIK"`` → ``"açik"``
    (noktalı i) ve ``"PASİF"`` → ``"pasi̇f"`` (i + birleşen nokta). İkisi de
    Türkçe sözlükte yok, yani büyük harf yazılmış bir "AÇIK/KAPALI" sütunu
    sessizce tanınmaz hale gelirdi. Önce ``I → ı`` ve ``İ → i`` eşlenir.
    """
    return str(text).replace("İ", "i").replace("I", "ı").casefold()


def parse_bool(value: Any) -> Optional[bool]:
    if isinstance(value, bool):
        return value
    if value is None:
        return None
    text = _tr_casefold(str(value).strip())
    if text in _TRUE_WORDS:
        return True
    if text in _FALSE_WORDS:
        return False
    return None


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip() or value.strip() in {"-", "—", "N/A", "n/a", "NULL", "null", "#YOK", "#N/A"}
    return False


# --- çağrı merkezi sözlüğü ---------------------------------------------------


@dataclass(frozen=True)
class MetricHint:
    """Bir sütun adının işaret ettiği çağrı merkezi göstergesi."""

    key: str
    label: str
    #: Hangi yön iyidir: ``up``, ``down`` ya da ``""`` (yorumsuz).
    better: str
    unit: str = ""
    #: Ölçüt mü (sayısal gösterge) yoksa kırılım mı (kategori/tarih)?
    dimension: bool = False


#: Sıra ÖNEMLİ: daha özel anahtar sözcük önce gelir, yoksa "çağrı süresi"
#: "çağrı" sayısına takılır.
METRIC_PATTERNS: Tuple[Tuple[str, MetricHint], ...] = (
    (r"\baht\b|ortalama\s*(g[oö]r[uü][sş]me|[cç]a[gğ]r[ıi]|i[sş]lem)|handle\s*time|talk\s*time|"
     r"g[oö]r[uü][sş]me\s*s[uü]re|konu[sş]ma\s*s[uü]re|[cç]a[gğ]r[ıi]\s*s[uü]re",
     MetricHint("aht", "Ortalama Görüşme Süresi (AHT)", "down", "sn")),
    (r"\basa\b|beklet?me\s*s[uü]re|bekleme|answer\s*speed|queue\s*time|kuyruk\s*s[uü]re|wait",
     MetricHint("asa", "Ortalama Cevaplama Süresi (ASA)", "down", "sn")),
    (r"\bacw\b|[cç]a[gğ]r[ıi]\s*sonras[ıi]|after\s*call",
     MetricHint("acw", "Çağrı Sonrası İş (ACW)", "down", "sn")),
    (r"\bhold\b|beklet(me)?\s*(s[uü]resi)?\b",
     MetricHint("hold", "Bekletme Süresi", "down", "sn")),
    (r"\bsla\b|servis\s*seviye|service\s*level|hizmet\s*seviye",
     MetricHint("sla", "Servis Seviyesi (SLA)", "up", "%")),
    (r"terk|abandon|vazge[cç]|kay[ıi]p\s*[cç]a[gğ]r[ıi]",
     MetricHint("abandon", "Terk Edilen Çağrı", "down")),
    (r"doluluk|occupancy|me[sş]guliyet|utilization",
     MetricHint("occupancy", "Doluluk", "up", "%")),
    (r"uyum|adherence|schedule\s*adher",
     MetricHint("adherence", "Vardiya Uyumu", "up", "%")),
    (r"csat|memnuniyet|satisfaction|puan|score|rating",
     MetricHint("csat", "Müşteri Memnuniyeti (CSAT)", "up")),
    (r"\bnps\b|tavsiye\s*skor",
     MetricHint("nps", "Net Tavsiye Skoru (NPS)", "up")),
    (r"\bfcr\b|ilk\s*(temas|[cç]a[gğ]r[ıi])|first\s*call\s*resolution|tek\s*seferde",
     MetricHint("fcr", "İlk Temasta Çözüm (FCR)", "up", "%")),
    (r"transfer|aktar[ıi]m|y[oö]nlendir",
     MetricHint("transfer", "Aktarım", "down")),
    (r"cevap(lanan|lama)|kar[sş][ıi]lanan|answered|handled|ba[gğ]lanan",
     MetricHint("answered", "Cevaplanan Çağrı", "up")),
    # "Çağrı Sebebi" / "Çağrı Konusu" bir SAYI değil bir KIRILIMDIR; genel
    # "çağrı" kalıbından ÖNCE gelmezse rapor bu sütunu "Çağrı Adedi" diye
    # etiketler ve tabloya bakan kişiye anlamsız bir başlık gösterir.
    (r"sebe[pb]|neden|reason|konu|topic|kategori|category|t[uü]r|type|sonu[cç]|disposition",
     MetricHint("reason", "Konu / Sebep", "", dimension=True)),
    (r"[cç]a[gğ]r[ıi]|\bcall\b|\bcalls\b|arama|contact|etkile[sş]im|i[sş]lem\s*adedi|hacim|volume|adet",
     MetricHint("calls", "Çağrı Adedi", "")),
    # Kırılımlar (ölçüt değil, eksen).
    (r"temsilci|agent|m[uü][sş]teri\s*temsilcisi|personel|dan[ıi][sş]man|operat[oö]r|[cç]al[ıi][sş]an",
     MetricHint("agent", "Temsilci", "", dimension=True)),
    (r"vardiya|shift|mesai",
     MetricHint("shift", "Vardiya", "", dimension=True)),
    (r"kuyruk|queue|beceri|skill|kampanya|proje",
     MetricHint("queue", "Kuyruk", "", dimension=True)),
    (r"kanal|channel|medya|ortam",
     MetricHint("channel", "Kanal", "", dimension=True)),
    (r"saat\s*dilim|interval|zaman\s*dilim|periy[oa]t|yar[ıi]m\s*saat|time\s*slot|\bsaat\b|\bhour\b",
     MetricHint("interval", "Saat Dilimi", "", dimension=True)),
    (r"tak[ıi]m|team|ekip|grup|group|departman|birim",
     MetricHint("team", "Takım", "", dimension=True)),
    (r"tarih|\bdate\b|\bg[uü]n\b|\bday\b|hafta|week|\bay\b|month",
     MetricHint("date", "Tarih", "", dimension=True)),
)

#: Yüzde biriminde olduğu ad ından anlaşılan sütunlar.
_PERCENT_NAME = re.compile(r"%|oran|rate|y[uü]zde|pct|percent", re.IGNORECASE)


def _normalise_name(name: str) -> str:
    text = str(name or "").strip().casefold()
    text = text.replace("i̇", "i")
    return re.sub(r"[_\-./]+", " ", text)


def detect_metric(name: str) -> Optional[MetricHint]:
    """Sütun adından çağrı merkezi göstergesini tanır (TR + EN)."""
    text = _normalise_name(name)
    if not text:
        return None
    for pattern, hint in METRIC_PATTERNS:
        if re.search(pattern, text, re.IGNORECASE):
            return hint
    return None


# --- sütun profili -----------------------------------------------------------


@dataclass
class ColumnProfile:
    """Tek bir sütun hakkında bilinen her şey — JSON'a çevrilebilir."""

    name: str
    index: int
    kind: str = KIND_EMPTY
    count: int = 0
    nulls: int = 0
    invalid: int = 0
    distinct: int = 0
    unit: str = ""
    metric: str = ""
    metric_label: str = ""
    better: str = ""
    is_dimension: bool = False

    # sayısal
    minimum: Optional[float] = None
    maximum: Optional[float] = None
    mean: Optional[float] = None
    median: Optional[float] = None
    stdev: Optional[float] = None
    total: Optional[float] = None
    p25: Optional[float] = None
    p75: Optional[float] = None
    p90: Optional[float] = None

    # kategorik
    top_values: List[Tuple[str, int]] = field(default_factory=list)

    # tarih
    first: Optional[datetime] = None
    last: Optional[datetime] = None
    granularity: str = ""

    @property
    def null_rate(self) -> float:
        total = self.count + self.nulls
        return (self.nulls / total) if total else 0.0

    @property
    def cardinality(self) -> float:
        """Benzersiz değerlerin dolu değerlere oranı (0–1)."""
        return (self.distinct / self.count) if self.count else 0.0

    @property
    def kind_label(self) -> str:
        return KIND_LABELS.get(self.kind, self.kind)

    @property
    def is_numeric(self) -> bool:
        return self.kind == KIND_NUMERIC

    @property
    def is_date(self) -> bool:
        return self.kind == KIND_DATE

    @property
    def is_groupable(self) -> bool:
        """Kırılım ekseni olmaya uygun mu?

        Kategorik ve makul sayıda ayrı değeri olan sütunlar. 500 farklı değer
        taşıyan bir "müşteri numarası" kırılım değil, kimliktir.
        """
        if self.kind in (KIND_CATEGORICAL, KIND_BOOLEAN):
            return 1 < self.distinct <= 60
        return False

    def as_dict(self) -> Dict[str, Any]:
        payload: Dict[str, Any] = {
            "name": self.name,
            "index": self.index,
            "kind": self.kind,
            "kind_label": self.kind_label,
            "count": self.count,
            "nulls": self.nulls,
            "null_rate": round(self.null_rate, 4),
            "invalid": self.invalid,
            "distinct": self.distinct,
            "cardinality": round(self.cardinality, 4),
            "unit": self.unit,
        }
        if self.metric:
            payload.update(
                {
                    "metric": self.metric,
                    "metric_label": self.metric_label,
                    "better": self.better,
                    "is_dimension": self.is_dimension,
                }
            )
        if self.kind == KIND_NUMERIC:
            payload["stats"] = {
                "min": self.minimum,
                "max": self.maximum,
                "mean": self.mean,
                "median": self.median,
                "stdev": self.stdev,
                "sum": self.total,
                "p25": self.p25,
                "p75": self.p75,
                "p90": self.p90,
            }
        if self.top_values:
            payload["top_values"] = [{"value": v, "count": c} for v, c in self.top_values]
        if self.kind == KIND_DATE:
            payload["period"] = {
                "first": self.first.isoformat() if self.first else "",
                "last": self.last.isoformat() if self.last else "",
                "granularity": self.granularity,
            }
        return payload


# --- veri kümesi -------------------------------------------------------------


@dataclass
class Dataset:
    """Profillenmiş bir tablo: başlıklar, tiplenmiş satırlar ve sütun profilleri."""

    name: str = "Veri"
    source: str = ""
    source_kind: str = ""
    columns: List[ColumnProfile] = field(default_factory=list)
    rows: List[List[Any]] = field(default_factory=list)
    notes: List[str] = field(default_factory=list)
    #: Kaynağa özgü kimlik (Drive dosya kimliği, dosya yolu…).
    source_id: str = ""

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def column_count(self) -> int:
        return len(self.columns)

    @property
    def headers(self) -> List[str]:
        return [c.name for c in self.columns]

    def column(self, name: str) -> Optional[ColumnProfile]:
        target = _normalise_name(name)
        for profile in self.columns:
            if _normalise_name(profile.name) == target:
                return profile
        return None

    def index_of(self, name: str) -> int:
        profile = self.column(name)
        return profile.index if profile else -1

    def values(self, name: str) -> List[Any]:
        """Bir sütunun tiplenmiş değerleri (boşlar dahil, sıra korunur)."""
        index = self.index_of(name)
        if index < 0:
            return []
        return [row[index] if index < len(row) else None for row in self.rows]

    def numeric_columns(self) -> List[ColumnProfile]:
        return [c for c in self.columns if c.is_numeric]

    def date_columns(self) -> List[ColumnProfile]:
        return [c for c in self.columns if c.is_date]

    def dimension_columns(self) -> List[ColumnProfile]:
        return [c for c in self.columns if c.is_groupable]

    def metric_columns(self) -> List[ColumnProfile]:
        """Çağrı merkezi göstergesi olarak tanınan SAYISAL sütunlar."""
        return [c for c in self.columns if c.metric and not c.is_dimension and c.is_numeric]

    def call_center_score(self) -> float:
        """Verinin çağrı merkezi verisine benzeme oranı (0–1)."""
        if not self.columns:
            return 0.0
        hits = sum(1 for c in self.columns if c.metric)
        return hits / len(self.columns)

    def subset(self, rows: Sequence[Sequence[Any]], name: str = "") -> "Dataset":
        """Satırların bir alt kümesiyle YENİ bir veri kümesi üretir.

        Sütunların tipi, birimi ve tanınan göstergesi KORUNUR (yeniden çıkarım
        yapılsa "%85" bilgisi kaybolurdu, çünkü değerler artık sayı); yalnızca
        istatistikler yeniden hesaplanır. Filtreli analiz — "son 3 ay",
        "yalnızca akşam vardiyası" — bunun üstünde çalışır.
        """
        clone = Dataset(
            name=name or self.name,
            source=self.source,
            source_kind=self.source_kind,
            source_id=self.source_id,
            columns=[],
            rows=[list(row) for row in rows],
            notes=list(self.notes),
        )
        for profile in self.columns:
            fresh = ColumnProfile(
                name=profile.name,
                index=profile.index,
                kind=profile.kind,
                unit=profile.unit,
                metric=profile.metric,
                metric_label=profile.metric_label,
                better=profile.better,
                is_dimension=profile.is_dimension,
            )
            _recompute(fresh, [row[profile.index] if profile.index < len(row) else None for row in clone.rows])
            clone.columns.append(fresh)
        return clone

    def content_hash(self) -> str:
        """Satır içeriğinin kararlı özeti — canlı senkronizasyon bunu karşılaştırır."""
        import hashlib

        digest = hashlib.sha256()
        digest.update("".join(self.headers).encode("utf-8"))
        for row in self.rows:
            digest.update(b"")
            digest.update("".join(_hash_cell(cell) for cell in row).encode("utf-8"))
        return digest.hexdigest()[:32]

    def profile(self) -> Dict[str, Any]:
        """Kararlı, JSON'a çevrilebilir profil sözlüğü."""
        return {
            "profile_version": PROFILE_VERSION,
            "name": self.name,
            "source": self.source,
            "source_kind": self.source_kind,
            "source_id": self.source_id,
            "row_count": self.row_count,
            "column_count": self.column_count,
            "call_center_score": round(self.call_center_score(), 3),
            "columns": [c.as_dict() for c in self.columns],
            "notes": list(self.notes),
        }

    def to_json(self) -> str:
        return json.dumps(self.profile(), ensure_ascii=False, indent=2)

    def summary_tr(self) -> str:
        """Tek cümlelik Türkçe özet — rapor kapaklarında kullanılır."""
        kinds: Dict[str, int] = {}
        for column in self.columns:
            kinds[column.kind_label] = kinds.get(column.kind_label, 0) + 1
        parts = ", ".join(f"{count} {label}" for label, count in kinds.items())
        return (
            f"{self.row_count:,} satır × {self.column_count} sütun".replace(",", ".")
            + (f" ({parts})" if parts else "")
        )


def _hash_cell(cell: Any) -> str:
    if cell is None:
        return ""
    if isinstance(cell, datetime):
        return cell.isoformat()
    if isinstance(cell, float):
        return repr(round(cell, 6))
    return str(cell)


# --- profilleme --------------------------------------------------------------


def _percentile(values: Sequence[float], fraction: float) -> Optional[float]:
    if not values:
        return None
    ordered = sorted(values)
    if len(ordered) == 1:
        return float(ordered[0])
    position = fraction * (len(ordered) - 1)
    low = int(math.floor(position))
    high = int(math.ceil(position))
    if low == high:
        return float(ordered[low])
    weight = position - low
    return float(ordered[low] * (1 - weight) + ordered[high] * weight)


def _granularity(moments: Sequence[datetime]) -> str:
    unique = sorted({m for m in moments})
    if len(unique) < 2:
        return "tek nokta"
    deltas = [
        (unique[i + 1] - unique[i]).total_seconds()
        for i in range(len(unique) - 1)
    ]
    deltas = [d for d in deltas if d > 0]
    if not deltas:
        return "tek nokta"
    step = statistics.median(deltas)
    if step < 120:
        return "dakika"
    if step < 3000:
        return f"{int(round(step / 60))} dakika"
    if step < 7200:
        return "saat"
    if step < 129600:  # 1.5 gün
        return "gün"
    if step < 950400:  # 11 gün
        return "hafta"
    if step < 3456000:  # 40 gün
        return "ay"
    return "uzun dönem"


def _looks_like_code(sample: Sequence[Any]) -> bool:
    """Sütun sayı DEĞİL kod mu taşıyor? (``0012``, ``05321234567``)

    Baştaki sıfırı olan bir dizi sayıya çevrilirse geri dönüşü olmayan biçimde
    bozulur. Yarıdan fazlası bu kalıba uyuyorsa sütun kategorik/metin sayılır ve
    değerler METİN olarak korunur.
    """
    if not sample:
        return False
    codes = sum(
        1 for v in sample if isinstance(v, str) and _LEADING_ZERO_CODE.match(v.strip())
    )
    return codes / len(sample) >= 0.5


def _pick_kind(raw: Sequence[Any], name: str) -> Tuple[str, str]:
    """Ham değerlerden sütun tipini ve birimini seçer."""
    sample = [v for v in raw[:TYPE_SAMPLE] if not is_blank(v)]
    if not sample:
        return KIND_EMPTY, ""

    booleans = sum(1 for v in sample if parse_bool(v) is not None)
    if booleans / len(sample) >= 0.95 and len({str(v).strip().casefold() for v in sample}) <= 4:
        return KIND_BOOLEAN, ""

    dates = sum(1 for v in sample if parse_datetime(v) is not None)
    if dates / len(sample) >= TYPE_THRESHOLD:
        return KIND_DATE, ""

    numbers = sum(1 for v in sample if parse_number(v) is not None)
    if numbers / len(sample) >= TYPE_THRESHOLD and not _looks_like_code(sample):
        unit = ""
        if any(isinstance(v, str) and "%" in v for v in sample) or _PERCENT_NAME.search(str(name or "")):
            unit = "%"
        elif any(isinstance(v, str) and _DURATION_RE.match(v.strip()) for v in sample):
            unit = "sn"
        elif any(isinstance(v, str) and any(sym in v for sym in _CURRENCY) for v in sample):
            unit = "₺"
        return KIND_NUMERIC, unit

    distinct = len({str(v).strip() for v in sample})
    avg_length = sum(len(str(v)) for v in sample) / len(sample)
    if avg_length > 45 or (distinct > 50 and distinct / len(sample) > 0.75):
        return KIND_TEXT, ""
    return KIND_CATEGORICAL, ""


def _convert(kind: str, value: Any) -> Any:
    if is_blank(value):
        return None
    if kind == KIND_NUMERIC:
        return parse_number(value)
    if kind == KIND_DATE:
        return parse_datetime(value)
    if kind == KIND_BOOLEAN:
        return parse_bool(value)
    return str(value).strip()


def _fill_stats(profile: ColumnProfile, values: Sequence[Any]) -> None:
    """Tiplenmiş değerlerden sayaçları ve istatistikleri doldurur.

    Hem ilk profilleme hem :meth:`Dataset.subset` buradan geçer, böylece
    filtrelenmiş bir veri kümesinin istatistikleri orijinaliyle aynı kurallarla
    hesaplanır.
    """
    seen: Dict[str, int] = {}
    numbers: List[float] = []
    moments: List[datetime] = []
    missing = 0
    profile.count = 0
    for value in values:
        if value is None:
            missing += 1
            continue
        profile.count += 1
        if profile.kind == KIND_NUMERIC and isinstance(value, (int, float)):
            numbers.append(float(value))
        elif profile.kind == KIND_DATE and isinstance(value, datetime):
            moments.append(value)
        key = _hash_cell(value)
        seen[key] = seen.get(key, 0) + 1

    profile.nulls = max(0, missing - profile.invalid)
    profile.distinct = len(seen)
    profile.minimum = profile.maximum = profile.mean = profile.median = None
    profile.stdev = profile.total = profile.p25 = profile.p75 = profile.p90 = None
    profile.first = profile.last = None
    profile.granularity = ""
    profile.top_values = []

    if profile.kind == KIND_NUMERIC and numbers:
        profile.minimum = min(numbers)
        profile.maximum = max(numbers)
        profile.mean = statistics.fmean(numbers)
        profile.median = statistics.median(numbers)
        profile.total = math.fsum(numbers)
        profile.stdev = statistics.pstdev(numbers) if len(numbers) > 1 else 0.0
        profile.p25 = _percentile(numbers, 0.25)
        profile.p75 = _percentile(numbers, 0.75)
        profile.p90 = _percentile(numbers, 0.90)
        if not profile.unit and profile.metric in ("sla", "occupancy", "adherence", "fcr"):
            if profile.maximum is not None and profile.maximum <= 100:
                profile.unit = "%"
    elif profile.kind == KIND_DATE and moments:
        profile.first = min(moments)
        profile.last = max(moments)
        profile.granularity = _granularity(moments)

    if profile.kind in (KIND_CATEGORICAL, KIND_BOOLEAN, KIND_TEXT):
        ordered = sorted(seen.items(), key=lambda item: (-item[1], item[0]))
        profile.top_values = [(label, count) for label, count in ordered[:TOP_VALUES]]


def _recompute(profile: ColumnProfile, values: Sequence[Any]) -> None:
    """Alt küme için istatistikleri sıfırdan hesaplar (tip ve birim korunur)."""
    profile.invalid = 0
    _fill_stats(profile, values)


def profile_table(
    headers: Sequence[Any],
    rows: Sequence[Sequence[Any]],
    name: str = "Veri",
    source: str = "",
    source_kind: str = "",
    source_id: str = "",
) -> Dataset:
    """Başlık + ham satırlardan profillenmiş bir :class:`Dataset` üretir.

    Bütün yükleyiciler buraya bağlanır, böylece profil kuralları TEK yerde
    tanımlıdır ve CSV ile Google E-Tablo aynı veriden aynı profili çıkarır.
    """
    clean_headers = _unique_headers(headers)
    if not clean_headers:
        raise DatasetError("Tabloda hiç sütun başlığı bulunamadı; ilk satırın başlık satırı olduğundan emin olun.")

    notes: List[str] = []
    body = [list(row) for row in rows]
    if len(body) > MAX_ROWS:
        notes.append(
            f"Veri {len(body):,} satır olduğu için ilk {MAX_ROWS:,} satır analiz edildi.".replace(",", ".")
        )
        body = body[:MAX_ROWS]

    width = len(clean_headers)
    normalised: List[List[Any]] = []
    for row in body:
        cells = list(row[:width])
        if len(cells) < width:
            cells += [None] * (width - len(cells))
        normalised.append(cells)

    # Tamamen boş satırlar (Excel dışa aktarımlarının klasik kuyruğu) atılır.
    normalised = [row for row in normalised if any(not is_blank(cell) for cell in row)]

    columns: List[ColumnProfile] = []
    typed_rows: List[List[Any]] = [[None] * width for _ in normalised]

    for index, header in enumerate(clean_headers):
        raw = [row[index] for row in normalised]
        kind, unit = _pick_kind(raw, header)
        hint = detect_metric(header)

        profile = ColumnProfile(name=header, index=index, kind=kind, unit=unit)
        if hint:
            profile.metric = hint.key
            profile.metric_label = hint.label
            profile.better = hint.better
            profile.is_dimension = hint.dimension
            if hint.unit and not profile.unit and kind == KIND_NUMERIC:
                # Ada göre bilinen birim, ancak sütun gerçekten sayısalsa ve
                # değerler yüzde aralığındaysa uygulanır.
                profile.unit = hint.unit if hint.unit != "%" else profile.unit or "%"

        typed: List[Any] = []
        for row_index, value in enumerate(raw):
            converted = _convert(kind, value)
            typed_rows[row_index][index] = converted
            typed.append(converted)
            if converted is None and not is_blank(value):
                profile.invalid += 1

        _fill_stats(profile, typed)
        columns.append(profile)

    dataset = Dataset(
        name=name or "Veri",
        source=source,
        source_kind=source_kind,
        source_id=source_id,
        columns=columns,
        rows=typed_rows,
        notes=notes,
    )

    if not dataset.rows:
        raise DatasetError("Veri kaynağında başlık dışında hiç satır yok; analiz edilecek bir şey bulunamadı.")

    empty = [c.name for c in columns if c.kind == KIND_EMPTY]
    if empty:
        dataset.notes.append("Tamamen boş sütun(lar) analiz dışı bırakıldı: " + ", ".join(empty[:8]) + ".")
    high_null = [c.name for c in columns if 0.4 < c.null_rate < 1.0]
    if high_null:
        dataset.notes.append(
            "Boşluk oranı %40'ın üstünde olan sütun(lar): " + ", ".join(high_null[:8]) + "."
        )
    return dataset


def _unique_headers(headers: Sequence[Any]) -> List[str]:
    """Başlıkları temizler, boş olanları adlandırır, tekrarları numaralandırır."""
    out: List[str] = []
    used: Dict[str, int] = {}
    trailing_blank = True
    cleaned: List[str] = []
    for raw in headers:
        text = "" if raw is None else str(raw).strip()
        cleaned.append(text)
    # Sağdaki boş başlık kuyruğu (Excel'in hayalet sütunları) atılır.
    while cleaned and not cleaned[-1]:
        cleaned.pop()
    for position, text in enumerate(cleaned, start=1):
        if not text:
            text = f"Sütun {position}"
        text = re.sub(r"\s+", " ", text)
        count = used.get(text.casefold(), 0)
        used[text.casefold()] = count + 1
        out.append(text if count == 0 else f"{text} ({count + 1})")
    del trailing_blank
    return out


# --- yükleyiciler ------------------------------------------------------------


def _sniff_delimiter(sample: str) -> str:
    candidates = [";", "\t", ",", "|"]
    first_lines = [line for line in sample.splitlines() if line.strip()][:10]
    if not first_lines:
        return ","
    best, best_score = ",", -1.0
    for candidate in candidates:
        counts = [line.count(candidate) for line in first_lines]
        if not counts or max(counts) == 0:
            continue
        # Tutarlılık ödüllendirilir: her satırda aynı sayıda görünen ayırıcı
        # gerçek ayırıcıdır, metnin içinde geçen virgül değil.
        spread = max(counts) - min(counts)
        score = statistics.fmean(counts) - spread
        if score > best_score:
            best, best_score = candidate, score
    return best


def load_text(
    text: str,
    name: str = "Yapıştırılan veri",
    delimiter: Optional[str] = None,
) -> Dataset:
    """Yapıştırılmış TSV/CSV metnini okur (ayırıcı otomatik saptanır)."""
    body = str(text or "").strip("﻿\n\r ")
    if not body.strip():
        raise DatasetError("Yapıştırılan metin boş; en az bir başlık satırı ve bir veri satırı gerekiyor.")
    sep = delimiter or _sniff_delimiter(body)
    try:
        reader = csv.reader(io.StringIO(body), delimiter=sep)
        table = [row for row in reader]
    except csv.Error as exc:
        raise DatasetError(f"Yapıştırılan metin tablo olarak okunamadı: {exc}") from exc
    table = [row for row in table if any(str(cell).strip() for cell in row)]
    if len(table) < 2:
        raise DatasetError(
            "Yapıştırılan metinde başlık satırının altında veri yok; en az iki satır gerekiyor."
        )
    return profile_table(table[0], table[1:], name=name, source="yapıştırılan metin", source_kind="text")


def load_csv(path: str, name: str = "", encoding: str = "") -> Dataset:
    """Bir ``.csv`` dosyasını okur. Kodlama verilmezse UTF-8 → CP1254 denenir."""
    if not os.path.exists(path):
        raise DatasetError(f"Dosya bulunamadı: {path}")
    encodings = [encoding] if encoding else ["utf-8-sig", "utf-8", "cp1254", "latin-1"]
    last_error: Optional[Exception] = None
    for candidate in encodings:
        try:
            with open(path, "r", encoding=candidate, newline="") as handle:
                body = handle.read()
            break
        except (UnicodeDecodeError, LookupError) as exc:
            last_error = exc
            continue
        except OSError as exc:
            raise DatasetError(f"Dosya okunamadı ({os.path.basename(path)}): {exc}") from exc
    else:
        raise DatasetError(
            f"Dosyanın karakter kodlaması çözülemedi ({os.path.basename(path)}): {last_error}"
        )

    if not body.strip():
        raise DatasetError(f"Dosya boş: {os.path.basename(path)}")
    sep = _sniff_delimiter(body)
    try:
        table = [row for row in csv.reader(io.StringIO(body), delimiter=sep)]
    except csv.Error as exc:
        raise DatasetError(f"CSV çözümlenemedi ({os.path.basename(path)}): {exc}") from exc
    table = [row for row in table if any(str(cell).strip() for cell in row)]
    if len(table) < 2:
        raise DatasetError(f"CSV dosyasında başlık dışında satır yok: {os.path.basename(path)}")
    return profile_table(
        table[0],
        table[1:],
        name=name or os.path.splitext(os.path.basename(path))[0],
        source=os.path.basename(path),
        source_kind="csv",
        source_id=os.path.abspath(path),
    )


def load_excel(path: str, sheet: str = "", name: str = "") -> Dataset:
    """Bir ``.xlsx``/``.xlsm`` dosyasının bir sayfasını okur.

    Sayfa verilmezse EN ÇOK DOLU HÜCREYE sahip sayfa seçilir: dışa aktarımların
    ilk sayfası çoğu zaman bir kapak ya da parametre sayfasıdır.
    """
    if not os.path.exists(path):
        raise DatasetError(f"Dosya bulunamadı: {path}")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - bağımlılık kurulu olmalı
        raise DatasetError(
            "Excel okuma kütüphanesi (openpyxl) kurulu değil. Kurulum: pip install openpyxl"
        ) from exc

    if path.lower().endswith(".xls"):
        raise DatasetError(
            "Eski ``.xls`` biçimi doğrudan okunamıyor. Dosyayı Excel'de açıp "
            "'.xlsx' ya da '.csv' olarak kaydedin."
        )

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise DatasetError(
            f"Excel dosyası açılamadı ({os.path.basename(path)}): parola korumalı ya da bozuk olabilir."
        ) from exc

    try:
        if sheet:
            if sheet not in workbook.sheetnames:
                raise DatasetError(
                    f"'{sheet}' sayfası bulunamadı. Dosyadaki sayfalar: "
                    + ", ".join(workbook.sheetnames)
                )
            worksheet = workbook[sheet]
        else:
            worksheet = _busiest_sheet(workbook)
        if worksheet is None:
            raise DatasetError(f"Excel dosyasında dolu bir sayfa yok: {os.path.basename(path)}")
        table = [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        try:
            workbook.close()
        except Exception:  # pragma: no cover - kapatma hatası raporu bozmaz
            pass

    table = [row for row in table if any(not is_blank(cell) for cell in row)]
    if len(table) < 2:
        raise DatasetError(
            f"'{worksheet.title}' sayfasında başlık dışında veri yok: {os.path.basename(path)}"
        )
    return profile_table(
        table[0],
        table[1:],
        name=name or worksheet.title,
        source=f"{os.path.basename(path)} · {worksheet.title}",
        source_kind="excel",
        source_id=os.path.abspath(path),
    )


def _busiest_sheet(workbook: Any) -> Any:
    best, best_size = None, -1
    for worksheet in workbook.worksheets:
        rows = worksheet.max_row or 0
        cols = worksheet.max_column or 0
        size = rows * cols
        if rows >= 2 and size > best_size:
            best, best_size = worksheet, size
    return best or (workbook.worksheets[0] if workbook.worksheets else None)


# --- Google E-Tablo ----------------------------------------------------------

_SHEET_ID_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9_\-]{20,})")
_GID_RE = re.compile(r"[#&?]gid=(\d+)")


def sheet_id_from_url(url: str) -> str:
    """Bir Google E-Tablo bağlantısından dosya kimliğini çıkarır."""
    text = str(url or "").strip()
    if not text:
        return ""
    match = _SHEET_ID_RE.search(text)
    if match:
        return match.group(1)
    if re.fullmatch(r"[a-zA-Z0-9_\-]{20,}", text):
        return text
    return ""


def load_google_sheet(
    url_or_id: str,
    sheet: str = "",
    name: str = "",
    service: Any = None,
) -> Dataset:
    """Bir Google E-Tabloyu mevcut Google kimlik doğrulamasıyla okur.

    ``service`` verilirse (testler ve çağıran taraf için) o kullanılır; yoksa
    ``google_auth.get_credentials()`` ile Sheets API v4 istemcisi kurulur.

    NOT — KAPSAM: depodaki ortak OAuth kapsamı ``drive.metadata.readonly``dir,
    yani e-tablo İÇERİĞİNİ okumaz. Bu çağrının çalışması için token'a
    ``https://www.googleapis.com/auth/spreadsheets.readonly`` kapsamının
    eklenmesi gerekir; eksikse aşağıdaki mesaj tam olarak bunu söyler.
    """
    file_id = sheet_id_from_url(url_or_id)
    if not file_id:
        raise DatasetError(
            "Google E-Tablo bağlantısı tanınmadı. Tarayıcıdaki tam adresi "
            "(https://docs.google.com/spreadsheets/d/... ) yapıştırın."
        )

    if service is None:
        service = _sheets_service()

    target = sheet or ""
    try:
        if not target:
            meta = service.spreadsheets().get(spreadsheetId=file_id, fields="properties.title,sheets.properties").execute()
            title = str((meta.get("properties") or {}).get("title") or "")
            sheets = [s.get("properties", {}) for s in (meta.get("sheets") or [])]
            visible = [s for s in sheets if not s.get("hidden")] or sheets
            if not visible:
                raise DatasetError("Google E-Tabloda okunabilir sayfa yok.")
            target = str(visible[0].get("title") or "")
        else:
            title = ""
        response = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=file_id, range=target, valueRenderOption="UNFORMATTED_VALUE")
            .execute()
        )
    except DatasetError:
        raise
    except Exception as exc:  # googleapiclient hataları dahil
        raise DatasetError(_google_error_message(exc)) from exc

    table = [list(row) for row in (response.get("values") or [])]
    table = [row for row in table if any(not is_blank(cell) for cell in row)]
    if len(table) < 2:
        raise DatasetError(
            f"Google E-Tablodaki '{target}' sayfasında başlık dışında veri yok."
        )
    return profile_table(
        table[0],
        table[1:],
        name=name or title or target or "Google E-Tablo",
        source=f"Google E-Tablo · {target}" if target else "Google E-Tablo",
        source_kind="google_sheet",
        source_id=file_id,
    )


def _sheets_service() -> Any:
    from ..integrations import google_auth

    if not google_auth.google_configured():
        raise DatasetError(
            "Google hesabı bağlı değil, e-tablo okunamıyor. "
            "Bağlamak için: python -m ai_assistant.integrations.google_auth"
        )
    try:
        creds = google_auth.get_credentials()
    except Exception as exc:
        raise DatasetError(f"Google kimlik doğrulaması başarısız: {exc}") from exc
    try:
        from googleapiclient.discovery import build

        return build("sheets", "v4", credentials=creds, cache_discovery=False)
    except ImportError as exc:
        raise DatasetError(
            "Google API istemcisi kurulu değil. Kurulum: pip install google-api-python-client"
        ) from exc
    except Exception as exc:
        raise DatasetError(f"Google Sheets servisi kurulamadı: {exc}") from exc


def _google_error_message(exc: Exception) -> str:
    text = str(exc)
    if "403" in text or "insufficient" in text.lower() or "scope" in text.lower():
        return (
            "Google E-Tablo okunamadı: mevcut yetki e-tablo içeriğini görmeye yetmiyor. "
            "Google bağlantısını 'spreadsheets.readonly' kapsamıyla yenileyin "
            "(python -m ai_assistant.integrations.google_auth)."
        )
    if "404" in text:
        return "Google E-Tablo bulunamadı; bağlantı yanlış olabilir ya da dosya bu hesapla paylaşılmamış."
    if "401" in text:
        return "Google oturumu geçersiz. Bağlantıyı yenileyin: python -m ai_assistant.integrations.google_auth"
    return f"Google E-Tablo okunamadı: {text}"


# --- tek kapı ----------------------------------------------------------------


def load_any(source: str, sheet: str = "", name: str = "") -> Dataset:
    """Ne verildiyse onu yükler: dosya yolu, e-tablo bağlantısı ya da ham metin.

    Kullanıcı "şu veriyi analiz et" derken hangi biçimi verdiğini söylemez; bu
    fonksiyon kararı verir.
    """
    text = str(source or "").strip()
    if not text:
        raise DatasetError("Analiz edilecek bir veri kaynağı verilmedi.")

    if text.startswith("http://") or text.startswith("https://"):
        if "docs.google.com/spreadsheets" in text or sheet_id_from_url(text):
            return load_google_sheet(text, sheet=sheet, name=name)
        raise DatasetError(
            "Bu bağlantı bir Google E-Tablo değil. Excel/CSV dosyasını indirip yolunu verin "
            "ya da e-tablo bağlantısını yapıştırın."
        )

    if "\n" in text or "\t" in text:
        return load_text(text, name=name or "Yapıştırılan veri")

    if os.path.exists(text):
        lower = text.lower()
        if lower.endswith((".xlsx", ".xlsm", ".xls")):
            return load_excel(text, sheet=sheet, name=name)
        if lower.endswith((".csv", ".txt", ".tsv")):
            return load_csv(text, name=name)
        # Uzantısız dosya: içeriğe bakılır.
        try:
            with open(text, "rb") as handle:
                head = handle.read(4)
        except OSError as exc:
            raise DatasetError(f"Dosya okunamadı: {exc}") from exc
        if head[:2] == b"PK":
            return load_excel(text, sheet=sheet, name=name)
        return load_csv(text, name=name)

    if len(text) >= 20 and re.fullmatch(r"[a-zA-Z0-9_\-]+", text):
        return load_google_sheet(text, sheet=sheet, name=name)

    raise DatasetError(
        "Veri kaynağı tanınmadı: dosya bulunamadı, geçerli bir e-tablo bağlantısı da değil. "
        "Excel/CSV dosya yolu, Google E-Tablo bağlantısı ya da tabloyu doğrudan yapıştırın."
    )


__all__ = [
    "ColumnProfile",
    "Dataset",
    "DatasetError",
    "KIND_BOOLEAN",
    "KIND_CATEGORICAL",
    "KIND_DATE",
    "KIND_EMPTY",
    "KIND_NUMERIC",
    "KIND_TEXT",
    "MetricHint",
    "PROFILE_VERSION",
    "detect_metric",
    "load_any",
    "load_csv",
    "load_excel",
    "load_google_sheet",
    "load_text",
    "parse_bool",
    "parse_datetime",
    "parse_number",
    "profile_table",
    "sheet_id_from_url",
]
