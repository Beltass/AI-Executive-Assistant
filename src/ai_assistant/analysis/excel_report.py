"""Gerçek bir ``.xlsx`` çalışma kitabı — resim değil, Excel'in kendi nesneleri.

Kullanıcının şikâyeti hep aynıydı: "çıktı amatör duruyor". Amatör duran şey
genelde şudur — ekran görüntüsü yapıştırılmış bir grafik, hizasız sütunlar,
1234.5678 diye yazılmış bir sayı, dondurulmamış başlık satırı.

Bu modül bunların hiçbirini yapmaz:

* Grafikler **yerel Excel grafikleridir** (:mod:`openpyxl.chart`). Kullanıcı
  tıklayınca veri aralığını görür, seriyi değiştirebilir, PowerPoint'e
  yapıştırınca canlı kalır.
* Sayılar **biçimlidir**: binlik ayırıcı, yüzdede tek ondalık, sürede
  ``dk:sn``, parada ``₺``.
* Başlık satırı **donmuş** ve **otomatik filtreli**, sütun genişlikleri içeriğe
  göre ayarlı, başlıklar koyu zeminde beyaz.
* KPI sütunlarında **koşullu biçimlendirme renk skalası** var: temsilci
  kırılımında kimin AHT'sinin uzadığı tabloya bakar bakmaz görülür.

Sayfa düzeni (hepsi Türkçe):

===================  ==========================================================
``Kapak``            Başlık, tarih, veri özeti, içindekiler, notlar
``Göstergeler``      KPI kartları: değer, taban, yön, Türkçe yorum
``Veri``             Ham veri (donmuş başlık + otomatik filtre)
``Kırılım – …``      Her kırılım için ayrı sayfa + çubuk/pasta grafik
``Zaman Serisi``     Dönemsel tablo + çizgi grafik (değer & hareketli ortalama)
``Aykırı Değerler``  IQR bandı dışındaki ölçümler ve yorumu
``İlişkiler``        Sayısal sütunlar arası korelasyon
===================  ==========================================================

``openpyxl`` kurulu değilse :class:`ai_assistant.analysis.dataset.DatasetError`
ile TÜRKÇE bir kurulum cümlesi döner — traceback değil.
"""

from __future__ import annotations

import logging
import os
import re
from datetime import date, datetime
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from .analyzer import (
    AGG_LABELS,
    AnalysisResult,
    Finding,
    PivotTable,
    TimeSeries,
    TONE_BAD,
    TONE_GOOD,
    TONE_WARN,
    format_duration,
    format_number,
    format_value,
)
from .dataset import (
    KIND_BOOLEAN,
    KIND_DATE,
    KIND_NUMERIC,
    Dataset,
    DatasetError,
)

logger = logging.getLogger(__name__)

# --- kurumsal palet (tek yerde, her sayfada aynı) ----------------------------

INK = "0F2537"  # başlık zemini — koyu lacivert
INK_SOFT = "1F4B63"  # ara başlık
ACCENT = "1F7A8C"  # vurgu
BAND = "F2F6F9"  # şerit dolgu
RULE = "D5DFE6"  # ince çizgi
GOOD_FILL = "DFF3E8"
GOOD_INK = "14624A"
WARN_FILL = "FDF0DA"
WARN_INK = "8A5300"
BAD_FILL = "FBE3E1"
BAD_INK = "9E241C"

TONE_STYLE = {
    TONE_GOOD: (GOOD_FILL, GOOD_INK, "İyi"),
    TONE_WARN: (WARN_FILL, WARN_INK, "Dikkat"),
    TONE_BAD: (BAD_FILL, BAD_INK, "Kritik"),
}

TREND_GLYPH = {"up": "▲", "down": "▼", "flat": "→"}

FMT_INT = "#,##0"
FMT_DEC = "#,##0.0"
FMT_PCT = '#,##0.0"%"'
FMT_SEC = '#,##0" sn"'
FMT_MONEY = '#,##0.00" ₺"'
FMT_DATE = "dd.mm.yyyy"
FMT_DATETIME = "dd.mm.yyyy hh:mm"

SHEET_COVER = "Kapak"
SHEET_KPI = "Göstergeler"
SHEET_DATA = "Veri"
SHEET_SERIES = "Zaman Serisi"
SHEET_OUTLIERS = "Aykırı Değerler"
SHEET_CORRELATION = "İlişkiler"

#: Ham veri sayfasına yazılan en fazla satır. Excel 1M satır kaldırır ama
#: 50 binden sonra dosya taşınamaz hale gelir; rapor bir dosya arşivi değil.
MAX_DATA_ROWS = 50_000

_INVALID_SHEET = re.compile(r"[\[\]:*?/\\]")


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise DatasetError(
            "Excel çıktısı için 'openpyxl' kütüphanesi gerekiyor ama kurulu değil. "
            "Kurulum: pip install openpyxl"
        ) from exc
    return __import__("openpyxl")


def _sheet_title(name: str, used: Iterable[str] = ()) -> str:
    """Excel'in kabul ettiği, 31 karakteri aşmayan, benzersiz sayfa adı."""
    clean = _INVALID_SHEET.sub(" ", str(name or "Sayfa")).strip() or "Sayfa"
    clean = re.sub(r"\s+", " ", clean)[:31]
    taken = {u.casefold() for u in used}
    if clean.casefold() not in taken:
        return clean
    for suffix in range(2, 60):
        candidate = f"{clean[:31 - len(str(suffix)) - 1]} {suffix}"
        if candidate.casefold() not in taken:
            return candidate
    return clean[:28] + "…"


def _number_format(unit: str, digits: Optional[int] = None) -> str:
    if unit == "%":
        return FMT_PCT
    if unit == "sn":
        return FMT_SEC
    if unit == "₺":
        return FMT_MONEY
    if digits == 0:
        return FMT_INT
    return FMT_DEC


# --- düşük seviyeli biçimlendirme -------------------------------------------


def _styles():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    return {
        "Alignment": Alignment,
        "Border": Border,
        "Font": Font,
        "PatternFill": PatternFill,
        "Side": Side,
    }


def _write_header(worksheet, row: int, labels: Sequence[str], width: int = 0) -> None:
    """Koyu zeminli, beyaz, ortalanmamış (sola dayalı) başlık satırı."""
    style = _styles()
    fill = style["PatternFill"]("solid", fgColor=INK)
    font = style["Font"](color="FFFFFF", bold=True, size=11)
    align = style["Alignment"](vertical="center", wrap_text=True)
    thin = style["Side"](style="thin", color=RULE)
    border = style["Border"](bottom=thin)
    for index, label in enumerate(labels, start=1):
        cell = worksheet.cell(row=row, column=index, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border
    worksheet.row_dimensions[row].height = 26 if width else 22


def _band_rows(worksheet, first_row: int, last_row: int, columns: int) -> None:
    """Tek satırlara açık şerit — uzun tabloda satır kaydırmayı önler."""
    style = _styles()
    fill = style["PatternFill"]("solid", fgColor=BAND)
    for row in range(first_row, last_row + 1):
        if (row - first_row) % 2 == 1:
            for column in range(1, columns + 1):
                worksheet.cell(row=row, column=column).fill = fill


def _autosize(worksheet, minimum: int = 10, maximum: int = 46) -> None:
    """Sütun genişliklerini içeriğe göre ayarlar."""
    from openpyxl.utils import get_column_letter

    widths: Dict[int, int] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value)
            if len(text) > 70:
                text = text[:70]
            widths[cell.column] = max(widths.get(cell.column, 0), len(text))
    for column, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column)].width = max(
            minimum, min(maximum, width + 3)
        )


def _title_block(
    worksheet, title: str, subtitle: str = "", row: int = 1, span: int = 6
) -> int:
    """Sayfanın üstündeki başlık bloğu; sonraki boş satırın numarasını verir."""
    style = _styles()
    worksheet.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=max(2, span)
    )
    cell = worksheet.cell(row=row, column=1, value=title)
    cell.font = style["Font"](size=16, bold=True, color=INK)
    cell.alignment = style["Alignment"](vertical="center")
    worksheet.row_dimensions[row].height = 28
    if subtitle:
        worksheet.merge_cells(
            start_row=row + 1, start_column=1, end_row=row + 1, end_column=max(2, span)
        )
        sub = worksheet.cell(row=row + 1, column=1, value=subtitle)
        sub.font = style["Font"](size=10, color=INK_SOFT)
        return row + 3
    return row + 2


# --- sayfalar ----------------------------------------------------------------


def _cover_sheet(
    worksheet, result: AnalysisResult, contents: Sequence[str], moment: datetime
) -> None:
    style = _styles()
    worksheet.sheet_view.showGridLines = False

    dataset = result.dataset
    row = _title_block(worksheet, result.title or "Veri Analiz Raporu", "", span=6)

    lead = worksheet.cell(row=row, column=1, value=result.headline)
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=6)
    lead.font = style["Font"](size=11, italic=True, color=INK_SOFT)
    lead.alignment = style["Alignment"](wrap_text=True, vertical="top")
    row += 3

    facts: List[Tuple[str, str]] = [
        ("Hazırlanma", moment.strftime("%d.%m.%Y %H:%M")),
        ("Veri kaynağı", dataset.source or dataset.name),
        ("Kapsam", dataset.summary_tr()),
        ("Satır sayısı", format_number(dataset.row_count, 0)),
        ("Sütun sayısı", format_number(dataset.column_count, 0)),
    ]
    date_columns = dataset.date_columns()
    if date_columns and date_columns[0].first and date_columns[0].last:
        facts.append(
            (
                "Dönem",
                f"{date_columns[0].first.strftime('%d.%m.%Y')} – "
                f"{date_columns[0].last.strftime('%d.%m.%Y')} "
                f"({date_columns[0].granularity} bazında)",
            )
        )
    facts.append(
        (
            "Veri tipi",
            (
                "Çağrı merkezi operasyon verisi"
                if result.is_call_center
                else "Genel tablo verisi"
            ),
        )
    )
    request = result.request
    if request is not None:
        if request.question:
            facts.append(("Sorulan soru", request.question))
        detail = request.summary_tr()
        if detail:
            facts.append(("Analiz tarifi", detail))

    _write_header(worksheet, row, ["Bilgi", "Değer"])
    row += 1
    first = row
    for label, value in facts:
        worksheet.cell(row=row, column=1, value=label).font = style["Font"](
            bold=True, color=INK_SOFT
        )
        worksheet.cell(row=row, column=2, value=value)
        row += 1
    _band_rows(worksheet, first, row - 1, 2)
    row += 1

    heading = worksheet.cell(row=row, column=1, value="İçindekiler")
    heading.font = style["Font"](size=12, bold=True, color=INK)
    row += 1
    for name in contents:
        cell = worksheet.cell(row=row, column=1, value=f"• {name}")
        cell.font = style["Font"](color=ACCENT)
        row += 1

    if result.notes:
        row += 1
        note_heading = worksheet.cell(row=row, column=1, value="Notlar")
        note_heading.font = style["Font"](size=12, bold=True, color=INK)
        row += 1
        for note in result.notes[:10]:
            cell = worksheet.cell(row=row, column=1, value=f"– {note}")
            cell.font = style["Font"](size=9, color=INK_SOFT)
            worksheet.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=6
            )
            row += 1

    worksheet.column_dimensions["A"].width = 26
    for letter in ("B", "C", "D", "E", "F"):
        worksheet.column_dimensions[letter].width = 18


def _kpi_sheet(workbook, result: AnalysisResult) -> None:
    style = _styles()
    worksheet = workbook.create_sheet(SHEET_KPI)
    worksheet.sheet_view.showGridLines = False
    row = _title_block(
        worksheet,
        "Öne Çıkan Göstergeler",
        "Her satır bir KPI: değeri, karşılaştırma tabanı ve yorumu",
        span=6,
    )

    headers = ["Gösterge", "Değer", "Karşılaştırma", "Fark", "Yön", "Durum", "Yorum"]
    _write_header(worksheet, row, headers)
    worksheet.freeze_panes = worksheet.cell(row=row + 1, column=1)
    row += 1
    first = row

    for finding in result.findings:
        fill, ink, label = TONE_STYLE.get(finding.tone, (None, INK_SOFT, "Bilgi"))
        worksheet.cell(row=row, column=1, value=finding.metric).font = style["Font"](
            bold=True, color=INK
        )
        value_cell = worksheet.cell(row=row, column=2, value=finding.display)
        value_cell.font = style["Font"](bold=True, size=12, color=ink)
        value_cell.alignment = style["Alignment"](horizontal="right")
        worksheet.cell(
            row=row,
            column=3,
            value=(
                f"{finding.baseline_label} ({format_value(finding.baseline, finding.unit)})"
                if finding.baseline is not None
                else "—"
            ),
        )
        worksheet.cell(
            row=row,
            column=4,
            value=(
                format_value(finding.delta, finding.unit)
                if finding.delta is not None
                else "—"
            ),
        ).alignment = style["Alignment"](horizontal="right")
        worksheet.cell(
            row=row, column=5, value=TREND_GLYPH.get(finding.direction, "→")
        ).alignment = style["Alignment"](horizontal="center")
        status = worksheet.cell(row=row, column=6, value=label)
        status.alignment = style["Alignment"](horizontal="center")
        if fill:
            status.fill = style["PatternFill"]("solid", fgColor=fill)
            status.font = style["Font"](bold=True, color=ink)
        comment = worksheet.cell(row=row, column=7, value=finding.interpretation)
        comment.alignment = style["Alignment"](wrap_text=True, vertical="top")
        worksheet.row_dimensions[row].height = 32
        row += 1

    if row == first:
        worksheet.cell(row=row, column=1, value="Bu veriden gösterge üretilemedi.")
    _autosize(worksheet)
    worksheet.column_dimensions["G"].width = 72


def _data_sheet(workbook, dataset: Dataset) -> None:
    from openpyxl.utils import get_column_letter

    worksheet = workbook.create_sheet(SHEET_DATA)
    _write_header(worksheet, 1, dataset.headers)
    worksheet.freeze_panes = "A2"

    formats: List[str] = []
    for column in dataset.columns:
        if column.kind == KIND_DATE:
            formats.append(
                FMT_DATETIME
                if (column.granularity or "").endswith(("dakika", "saat"))
                else FMT_DATE
            )
        elif column.kind == KIND_NUMERIC:
            digits = (
                0
                if column.median is not None
                and abs(column.median - round(column.median)) < 1e-9
                else 1
            )
            formats.append(_number_format(column.unit, digits))
        else:
            formats.append("")

    rows = dataset.rows[:MAX_DATA_ROWS]
    for offset, row in enumerate(rows, start=2):
        for index, value in enumerate(row, start=1):
            if isinstance(value, bool):
                value = "Evet" if value else "Hayır"
            cell = worksheet.cell(row=offset, column=index, value=value)
            fmt = formats[index - 1] if index - 1 < len(formats) else ""
            if fmt:
                cell.number_format = fmt

    last_row = len(rows) + 1
    if last_row >= 2:
        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(max(1, dataset.column_count))}{last_row}"
        )
    _autosize(worksheet, maximum=32)
    if len(dataset.rows) > MAX_DATA_ROWS:
        worksheet.cell(
            row=last_row + 2,
            column=1,
            value=f"Not: {format_number(len(dataset.rows), 0)} satırdan ilk "
            f"{format_number(MAX_DATA_ROWS, 0)} tanesi yazıldı.",
        )


def _pivot_sheet(
    workbook, pivot: PivotTable, used: Sequence[str], chart: str = "bar"
) -> str:
    """Bir kırılım tablosu + yerel Excel grafiği."""
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    style = _styles()
    axis = " × ".join(pivot.dimensions) if pivot.dimensions else "Özet"
    name = _sheet_title(f"Kırılım – {axis}", used)
    worksheet = workbook.create_sheet(name)
    worksheet.sheet_view.showGridLines = False

    row = _title_block(
        worksheet, pivot.title, pivot.note, span=len(pivot.column_keys) or 3
    )
    headers = list(pivot.dimensions) + [m.title() for m in pivot.measures]
    _write_header(worksheet, row, headers)
    header_row = row
    worksheet.freeze_panes = worksheet.cell(row=row + 1, column=1)
    row += 1
    first_data = row

    for entry in pivot.rows:
        for index, dimension in enumerate(pivot.dimensions, start=1):
            worksheet.cell(row=row, column=index, value=entry.get(dimension, ""))
        for offset, measure in enumerate(pivot.measures):
            column = len(pivot.dimensions) + offset + 1
            cell = worksheet.cell(row=row, column=column, value=entry.get(measure.key))
            digits = 1 if measure.unit in ("%",) else 0
            cell.number_format = _number_format(measure.unit, digits)
        row += 1
    last_data = row - 1
    _band_rows(worksheet, first_data, last_data, len(headers))

    # Toplam satırı — kalın üst çizgi, koyu yazı.
    if pivot.totals:
        thin = style["Side"](style="medium", color=INK)
        for index, dimension in enumerate(pivot.dimensions, start=1):
            cell = worksheet.cell(
                row=row,
                column=index,
                value=pivot.totals.get(dimension, "" if index > 1 else "TOPLAM"),
            )
            cell.font = style["Font"](bold=True, color=INK)
            cell.border = style["Border"](top=thin)
        for offset, measure in enumerate(pivot.measures):
            column = len(pivot.dimensions) + offset + 1
            cell = worksheet.cell(
                row=row, column=column, value=pivot.totals.get(measure.key)
            )
            cell.number_format = _number_format(
                measure.unit, 1 if measure.unit == "%" else 0
            )
            cell.font = style["Font"](bold=True, color=INK)
            cell.border = style["Border"](top=thin)
        total_row = row
        row += 1
    else:
        total_row = 0

    # Koşullu biçimlendirme: her ölçüt sütununa üç renkli skala.
    if last_data >= first_data:
        for offset, measure in enumerate(pivot.measures):
            letter = get_column_letter(len(pivot.dimensions) + offset + 1)
            worksheet.conditional_formatting.add(
                f"{letter}{first_data}:{letter}{last_data}",
                ColorScaleRule(
                    start_type="min",
                    start_color="FFF6E6EA" if measure.unit else "FFF3F7FA",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFFFF3D6",
                    end_type="max",
                    end_color="FFD8EFE4",
                ),
            )

    _autosize(worksheet)

    # Grafik: ilk ölçüt, ilk 12 kırılım.
    if pivot.rows and pivot.measures:
        chart_rows = min(last_data, first_data + 11)
        labels = Reference(worksheet, min_col=1, min_row=first_data, max_row=chart_rows)
        values = Reference(
            worksheet,
            min_col=len(pivot.dimensions) + 1,
            min_row=header_row,
            max_row=chart_rows,
        )
        if chart == "pie":
            figure = PieChart()
            figure.height = 8
        else:
            figure = BarChart()
            figure.type = "col"
            figure.gapWidth = 60
            figure.height = 8.5
            figure.y_axis.title = pivot.measures[0].title()
            figure.y_axis.majorGridlines = None
        figure.add_data(values, titles_from_data=True)
        figure.set_categories(labels)
        figure.title = f"{pivot.dimensions[0]} bazında {pivot.measures[0].title()}"
        figure.width = 20
        figure.style = 2
        anchor_row = (total_row or last_data) + 2
        worksheet.add_chart(figure, f"A{anchor_row}")
    return name


def _series_sheet(workbook, series_list: Sequence[TimeSeries]) -> None:
    from openpyxl.chart import LineChart, Reference

    style = _styles()
    worksheet = workbook.create_sheet(SHEET_SERIES)
    worksheet.sheet_view.showGridLines = False
    row = _title_block(
        worksheet,
        "Zaman Serisi",
        "Dönemsel değer, bir önceki döneme fark ve hareketli ortalama",
        span=5,
    )

    for series in series_list:
        block = worksheet.cell(row=row, column=1, value=series.title)
        block.font = style["Font"](size=12, bold=True, color=INK)
        row += 1
        trend_word = {"up": "yükseliş", "down": "düşüş", "flat": "yatay seyir"}.get(
            series.trend, ""
        )
        change = (
            f"{series.period} bazında {len(series.points)} dönem · "
            f"eğilim: {trend_word}"
            + (
                f" (%{format_number(abs(series.change_pct), 1)})"
                if series.change_pct is not None
                else ""
            )
        )
        note = worksheet.cell(row=row, column=1, value=change)
        note.font = style["Font"](size=9, color=INK_SOFT)
        row += 1

        _write_header(
            worksheet,
            row,
            ["Dönem", series.title, "Fark", "Değişim %", "Hareketli ort."],
        )
        header_row = row
        row += 1
        first = row
        digits = 1 if series.unit == "%" else 0
        for point in series.points:
            worksheet.cell(row=row, column=1, value=point.label)
            worksheet.cell(row=row, column=2, value=point.value).number_format = (
                _number_format(series.unit, digits)
            )
            worksheet.cell(row=row, column=3, value=point.delta).number_format = (
                _number_format(series.unit, digits)
            )
            pct = worksheet.cell(row=row, column=4, value=point.delta_pct)
            pct.number_format = FMT_PCT
            worksheet.cell(row=row, column=5, value=point.moving).number_format = (
                _number_format(series.unit, 1)
            )
            row += 1
        last = row - 1
        _band_rows(worksheet, first, last, 5)

        if last >= first:
            figure = LineChart()
            figure.title = series.title
            figure.style = 2
            figure.height = 8
            figure.width = 20
            figure.y_axis.title = series.unit or "Değer"
            figure.x_axis.title = f"Dönem ({series.period})"
            values = Reference(worksheet, min_col=2, min_row=header_row, max_row=last)
            moving = Reference(worksheet, min_col=5, min_row=header_row, max_row=last)
            figure.add_data(values, titles_from_data=True)
            figure.add_data(moving, titles_from_data=True)
            figure.set_categories(
                Reference(worksheet, min_col=1, min_row=first, max_row=last)
            )
            for line in figure.series:
                line.smooth = False
            worksheet.add_chart(figure, f"G{header_row}")
        row = max(row + 2, header_row + 18)

    _autosize(worksheet, maximum=28)


def _outlier_sheet(workbook, reports: Sequence[Any]) -> None:
    style = _styles()
    worksheet = workbook.create_sheet(SHEET_OUTLIERS)
    worksheet.sheet_view.showGridLines = False
    row = _title_block(
        worksheet,
        "Aykırı Değerler",
        "IQR (çeyrekler açıklığı) yöntemi: Q1−1,5×IQR ile Q3+1,5×IQR bandının dışı",
        span=6,
    )

    _write_header(
        worksheet,
        row,
        ["Ölçüt", "Alt sınır", "Üst sınır", "Aykırı adet", "Oran", "Yorum"],
    )
    worksheet.freeze_panes = worksheet.cell(row=row + 1, column=1)
    row += 1
    first = row
    for report in reports:
        worksheet.cell(row=row, column=1, value=report.column).font = style["Font"](
            bold=True, color=INK
        )
        worksheet.cell(row=row, column=2, value=report.lower).number_format = (
            _number_format(report.unit, 1)
        )
        worksheet.cell(row=row, column=3, value=report.upper).number_format = (
            _number_format(report.unit, 1)
        )
        worksheet.cell(row=row, column=4, value=report.count).number_format = FMT_INT
        worksheet.cell(row=row, column=5, value=report.rate * 100).number_format = (
            FMT_PCT
        )
        comment = worksheet.cell(row=row, column=6, value=report.interpretation)
        comment.alignment = style["Alignment"](wrap_text=True, vertical="top")
        worksheet.row_dimensions[row].height = 30
        row += 1
    _band_rows(worksheet, first, row - 1, 6)

    examples = [(r, o) for r in reports for o in r.examples]
    if examples:
        row += 1
        heading = worksheet.cell(row=row, column=1, value="En uç ölçümler")
        heading.font = style["Font"](size=12, bold=True, color=INK)
        row += 1
        _write_header(worksheet, row, ["Ölçüt", "Değer", "Bağlam", "Taraf", "Satır"])
        row += 1
        start = row
        for report, outlier in examples[:40]:
            worksheet.cell(row=row, column=1, value=report.column)
            worksheet.cell(row=row, column=2, value=outlier.value).number_format = (
                _number_format(report.unit, 1)
            )
            worksheet.cell(row=row, column=3, value=outlier.context or "—")
            worksheet.cell(row=row, column=4, value=f"{outlier.side} sınır")
            worksheet.cell(
                row=row, column=5, value=outlier.row_index + 2
            ).number_format = FMT_INT
            row += 1
        _band_rows(worksheet, start, row - 1, 5)
    _autosize(worksheet)
    worksheet.column_dimensions["F"].width = 74


def _correlation_sheet(workbook, correlations: Sequence[Any]) -> None:
    style = _styles()
    worksheet = workbook.create_sheet(SHEET_CORRELATION)
    worksheet.sheet_view.showGridLines = False
    row = _title_block(
        worksheet,
        "Göstergeler Arası İlişkiler",
        "Pearson korelasyon katsayısı (−1 ile +1 arası). İlişki nedensellik demek değildir.",
        span=5,
    )
    _write_header(worksheet, row, ["Gösterge 1", "Gösterge 2", "r", "Güç", "Yorum"])
    row += 1
    first = row
    for item in correlations:
        worksheet.cell(row=row, column=1, value=item.left)
        worksheet.cell(row=row, column=2, value=item.right)
        cell = worksheet.cell(row=row, column=3, value=round(item.r, 3))
        cell.number_format = "0.000"
        worksheet.cell(row=row, column=4, value=item.strength)
        comment = worksheet.cell(row=row, column=5, value=item.interpretation)
        comment.alignment = style["Alignment"](wrap_text=True, vertical="top")
        worksheet.row_dimensions[row].height = 28
        row += 1
    _band_rows(worksheet, first, row - 1, 5)

    if first <= row - 1:
        from openpyxl.formatting.rule import ColorScaleRule

        worksheet.conditional_formatting.add(
            f"C{first}:C{row - 1}",
            ColorScaleRule(
                start_type="num",
                start_value=-1,
                start_color="FFF6D6D2",
                mid_type="num",
                mid_value=0,
                mid_color="FFFFFFFF",
                end_type="num",
                end_value=1,
                end_color="FFD8EFE4",
            ),
        )
    _autosize(worksheet)
    worksheet.column_dimensions["E"].width = 80


# --- ana giriş ---------------------------------------------------------------


def build_workbook(
    result: AnalysisResult,
    path: str,
    moment: Optional[datetime] = None,
) -> str:
    """Analiz sonucunu bir ``.xlsx`` dosyasına yazar ve yolunu döner.

    Var olan dosyanın üzerine yazar. Yazılamıyorsa (izin, dolu disk)
    :class:`DatasetError` ile Türkçe bir mesaj verir.
    """
    _require_openpyxl()
    from openpyxl import Workbook

    when = moment or datetime.now()
    dataset = result.dataset
    workbook = Workbook()
    # Varsayılan sayfa atılır: kapak, içindekiler kesinleşsin diye EN SON ve
    # en başa eklenir.
    workbook.remove(workbook.active)

    contents: List[str] = [SHEET_KPI, SHEET_DATA]
    _kpi_sheet(workbook, result)
    _data_sheet(workbook, dataset)

    used: List[str] = [SHEET_COVER, SHEET_KPI, SHEET_DATA]
    for index, pivot in enumerate(result.pivots):
        if not pivot.rows:
            continue
        # İlk kırılım pasta olarak da anlamlıysa (az sayıda dilim) pasta çizilir.
        chart = (
            "pie"
            if len(pivot.rows) <= 6 and len(pivot.dimensions) == 1 and index == 0
            else "bar"
        )
        name = _pivot_sheet(workbook, pivot, used, chart=chart)
        used.append(name)
        contents.append(name)

    if result.series:
        _series_sheet(workbook, result.series)
        contents.append(SHEET_SERIES)
    if result.outliers:
        _outlier_sheet(workbook, result.outliers)
        contents.append(SHEET_OUTLIERS)
    if result.correlations:
        _correlation_sheet(workbook, result.correlations)
        contents.append(SHEET_CORRELATION)

    # Kapak en son yazılır: içindekiler artık gerçek sayfa adlarını gösteriyor.
    cover = workbook.create_sheet(SHEET_COVER, 0)
    _cover_sheet(cover, result, contents, when)
    workbook.active = 0

    directory = os.path.dirname(os.path.abspath(path))
    if directory:
        os.makedirs(directory, exist_ok=True)
    try:
        workbook.save(path)
    except OSError as exc:
        raise DatasetError(f"Excel dosyası kaydedilemedi ({path}): {exc}") from exc
    return path


__all__ = [
    "FMT_DATE",
    "FMT_INT",
    "FMT_PCT",
    "FMT_SEC",
    "SHEET_COVER",
    "SHEET_CORRELATION",
    "SHEET_DATA",
    "SHEET_KPI",
    "SHEET_OUTLIERS",
    "SHEET_SERIES",
    "build_workbook",
]
